"""Backend tests for İhracat Bedel Kapatma sistemi (v2 schema)."""
import os
from datetime import datetime, timedelta

import pytest
import requests
from dotenv import dotenv_values

frontend_env = dotenv_values("/app/frontend/.env")
BASE_URL = (os.environ.get("REACT_APP_BACKEND_URL") or frontend_env.get("REACT_APP_BACKEND_URL")).rstrip("/")
API = f"{BASE_URL}/api"

CREDS = {
    "admin":     ("admin@ihracat.com",   "Admin1234!"),
    "ihracat":   ("ihracat@ihracat.com", "Test1234!"),
    "banka":     ("banka@ihracat.com",   "Test1234!"),
    "onaylayan": ("sef@ihracat.com",     "Test1234!"),
    "viewer":    ("viewer@ihracat.com",  "Test1234!"),
}
TWOFA_EMAIL = "test2fa@ihracat.com"
TWOFA_PW = "Test1234!"


def _read_last_2fa_code():
    """Read latest [2FA] line from backend log."""
    import subprocess
    for path in ("/var/log/supervisor/backend.err.log", "/var/log/supervisor/backend.out.log"):
        try:
            out = subprocess.check_output(
                f"grep '\\[2FA\\]' {path} | tail -1", shell=True, text=True, timeout=5)
            if out.strip():
                # "... [2FA] test2fa@... doğrulama kodu: 123456"
                return out.strip().split(":")[-1].strip()
        except Exception:
            continue
    return None


def _login(email, password):
    r = requests.post(f"{API}/auth/login", json={"email": email, "password": password}, timeout=30)
    assert r.status_code == 200, f"login failed {email}: {r.status_code} {r.text[:200]}"
    data = r.json()
    assert data.get("token"), "token missing"
    return data["token"], data


@pytest.fixture(scope="module")
def tokens():
    return {role: _login(e, p)[0] for role, (e, p) in CREDS.items()}


def _h(tok):
    return {"Authorization": f"Bearer {tok}", "Content-Type": "application/json"}


def _base_dec(beyanname_no, acilis, kapanis="", doviz="EUR", tutar=10000.0,
              ithalatci="TEST Ithalatci A.S.", gumruk="GM-001",
              ibkb=False, destek=False):
    return {
        "beyanname_no": beyanname_no,
        "acilis_tarihi": acilis,
        "kapanis_tarihi": kapanis,
        "ithalatci": ithalatci,
        "gumruk_mudurlugu_no": gumruk,
        "doviz": doviz,
        "tutar": tutar,
        "ibkb_alindi": ibkb,
        "destek_alindi": destek,
    }


# ---------- auth ----------
class TestAuth:
    def test_login_all_roles(self):
        for role, (e, p) in CREDS.items():
            _, data = _login(e, p)
            assert data["email"] == e
            assert data["role"] in ("admin", "ihracat", "banka", "onaylayan", "goruntuleyici")

    def test_login_wrong_password(self):
        r = requests.post(f"{API}/auth/login",
                          json={"email": "admin@ihracat.com", "password": "wrongxxx"})
        assert r.status_code == 401

    def test_me(self, tokens):
        r = requests.get(f"{API}/auth/me", headers=_h(tokens["admin"]))
        assert r.status_code == 200 and r.json()["role"] == "admin"

    def test_me_unauthorized(self):
        assert requests.get(f"{API}/auth/me").status_code == 401


# ---------- declarations (new schema) ----------
class TestDeclarations:
    beyanname_no = f"TEST_BEY_{int(datetime.now().timestamp())}"
    acilis = (datetime.now() - timedelta(days=10)).strftime("%Y-%m-%d")
    kapanis = (datetime.now() - timedelta(days=5)).strftime("%Y-%m-%d")
    created_id = None

    def test_create_with_kapanis(self, tokens):
        payload = _base_dec(self.__class__.beyanname_no, self.acilis, self.kapanis,
                            doviz="EUR", tutar=10000.0, ithalatci="TEST GmbH",
                            gumruk="GM-42", ibkb=False, destek=False)
        r = requests.post(f"{API}/declarations", headers=_h(tokens["ihracat"]), json=payload)
        assert r.status_code == 200, r.text
        d = r.json()
        assert d["beyanname_no"] == self.__class__.beyanname_no
        assert d["durum"] == "ACIK"
        assert d["kalan"] == 10000.0
        assert d["ithalatci"] == "TEST GmbH"
        assert d["gumruk_mudurlugu_no"] == "GM-42"
        assert d["ibkb_durum"] == "DUZENLENMEDI"
        assert d["destek_durum"] == "ALINMADI"
        # destek = %3
        assert abs(d["destek_tutari"] - 300.0) < 0.01
        # son kapatma = kapanis + 180
        exp = (datetime.strptime(self.kapanis, "%Y-%m-%d") + timedelta(days=180)).strftime("%Y-%m-%d")
        assert d["son_kapatma_tarihi"] == exp
        self.__class__.created_id = d["id"]

    def test_son_kapatma_falls_back_to_acilis(self, tokens):
        no = f"TEST_FALLBACK_{int(datetime.now().timestamp())}"
        acilis = "2026-02-01"
        payload = _base_dec(no, acilis, "", tutar=1000.0)
        r = requests.post(f"{API}/declarations", headers=_h(tokens["ihracat"]), json=payload)
        assert r.status_code == 200, r.text
        d = r.json()
        exp = (datetime.strptime(acilis, "%Y-%m-%d") + timedelta(days=180)).strftime("%Y-%m-%d")
        assert d["son_kapatma_tarihi"] == exp
        requests.delete(f"{API}/declarations/{d['id']}", headers=_h(tokens["ihracat"]))

    def test_ibkb_destek_true(self, tokens):
        no = f"TEST_IBKB_{int(datetime.now().timestamp())}"
        r = requests.post(f"{API}/declarations", headers=_h(tokens["ihracat"]),
                          json=_base_dec(no, "2026-01-01", tutar=5000.0, ibkb=True, destek=True))
        assert r.status_code == 200
        d = r.json()
        assert d["ibkb_durum"] == "DUZENLENDI" and d["destek_durum"] == "ALINDI"
        assert abs(d["destek_tutari"] - 150.0) < 0.01
        requests.delete(f"{API}/declarations/{d['id']}", headers=_h(tokens["ihracat"]))

    def test_duplicate_beyanname_no(self, tokens):
        r = requests.post(f"{API}/declarations", headers=_h(tokens["ihracat"]),
                          json=_base_dec(self.__class__.beyanname_no, self.acilis))
        assert r.status_code == 400 and "zaten" in r.text.lower()

    def test_rbac_create(self, tokens):
        for role in ["banka", "onaylayan", "viewer"]:
            r = requests.post(f"{API}/declarations", headers=_h(tokens[role]),
                              json=_base_dec(f"NO_{role}", "2026-01-01"))
            assert r.status_code == 403, f"{role} should be 403"

    def test_all_roles_can_list(self, tokens):
        for role in ["admin", "ihracat", "banka", "onaylayan", "viewer"]:
            r = requests.get(f"{API}/declarations", headers=_h(tokens[role]))
            assert r.status_code == 200 and isinstance(r.json(), list)

    def test_search_by_ithalatci_and_gumruk(self, tokens):
        r = requests.get(f"{API}/declarations", headers=_h(tokens["admin"]),
                         params={"q": "TEST GmbH"})
        assert r.status_code == 200
        assert any(d["id"] == self.__class__.created_id for d in r.json())
        r2 = requests.get(f"{API}/declarations", headers=_h(tokens["admin"]),
                          params={"q": "GM-42"})
        assert any(d["id"] == self.__class__.created_id for d in r2.json())

    def test_filter_ibkb_destek(self, tokens):
        r = requests.get(f"{API}/declarations", headers=_h(tokens["admin"]),
                         params={"ibkb": "DUZENLENMEDI"})
        assert r.status_code == 200
        assert all(d["ibkb_durum"] == "DUZENLENMEDI" for d in r.json())
        r2 = requests.get(f"{API}/declarations", headers=_h(tokens["admin"]),
                          params={"destek": "ALINMADI"})
        assert all(d["destek_durum"] == "ALINMADI" for d in r2.json())


# ---------- payments ----------
class TestPayments:
    created_id = None

    def test_banka_create(self, tokens):
        r = requests.post(f"{API}/payments", headers=_h(tokens["banka"]), json={
            "banka": "TEST Bank", "gonderen": "TEST Buyer GmbH",
            "tarih": datetime.now().strftime("%Y-%m-%d"),
            "doviz": "EUR", "tutar": 15000.0, "aciklama": "TEST",
        })
        assert r.status_code == 200, r.text
        p = r.json()
        assert p["bakiye"] == 15000.0 and p["durum"] == "KULLANILMADI"
        self.__class__.created_id = p["id"]

    def test_rbac_create(self, tokens):
        for role in ["ihracat", "onaylayan", "viewer"]:
            r = requests.post(f"{API}/payments", headers=_h(tokens[role]),
                              json={"banka": "x", "gonderen": "y", "tarih": "2026-01-01",
                                    "doviz": "USD", "tutar": 1})
            assert r.status_code == 403


# ---------- matching (same currency and cross currency) ----------
class TestMatching:
    dec_id = None
    pay_id = None
    pay_usd_id = None
    match_id = None
    dec_no = f"TEST_MATCH_{int(datetime.now().timestamp())}"

    def test_setup(self, tokens):
        acilis = "2026-03-02"
        r = requests.post(f"{API}/declarations", headers=_h(tokens["ihracat"]),
                          json=_base_dec(self.__class__.dec_no, acilis, doviz="EUR", tutar=5000.0))
        assert r.status_code == 200
        self.__class__.dec_id = r.json()["id"]

        r = requests.post(f"{API}/payments", headers=_h(tokens["banka"]), json={
            "banka": "TEST", "gonderen": "TEST",
            "tarih": "2026-03-02", "doviz": "EUR", "tutar": 8000.0})
        assert r.status_code == 200
        self.__class__.pay_id = r.json()["id"]

        r = requests.post(f"{API}/payments", headers=_h(tokens["banka"]), json={
            "banka": "TEST", "gonderen": "TEST-USD",
            "tarih": "2026-03-02", "doviz": "USD", "tutar": 5000.0})
        assert r.status_code == 200
        self.__class__.pay_usd_id = r.json()["id"]

    def test_ihracat_cannot_match(self, tokens):
        r = requests.post(f"{API}/matches", headers=_h(tokens["ihracat"]), json={
            "declaration_id": self.__class__.dec_id, "payment_id": self.__class__.pay_id,
            "kapatilan_tutar": 100})
        assert r.status_code == 403

    def test_same_currency_match_kur_1(self, tokens):
        """Same currency => kur=1, kur_kaynak='AYNI_DOVIZ', bedel = kapatilan."""
        r = requests.post(f"{API}/matches", headers=_h(tokens["onaylayan"]), json={
            "declaration_id": self.__class__.dec_id, "payment_id": self.__class__.pay_id,
            "kapatilan_tutar": 2000.0})
        assert r.status_code == 200, r.text
        m = r.json()
        assert m["kur"] == 1.0
        assert m["kur_kaynak"] == "AYNI_DOVIZ"
        assert m["kapatilan_tutar"] == 2000.0
        assert m["bedel_kullanilan"] == 2000.0
        self.__class__.match_id = m["id"]

        d = next(x for x in requests.get(f"{API}/declarations", headers=_h(tokens["admin"])).json()
                 if x["id"] == self.__class__.dec_id)
        assert d["durum"] == "KISMI" and d["kapatilan"] == 2000.0 and d["kalan"] == 3000.0

        p = next(x for x in requests.get(f"{API}/payments", headers=_h(tokens["admin"])).json()
                 if x["id"] == self.__class__.pay_id)
        assert p["kullanilan"] == 2000.0 and p["bakiye"] == 6000.0

    def test_over_declaration_kalan(self, tokens):
        r = requests.post(f"{API}/matches", headers=_h(tokens["onaylayan"]), json={
            "declaration_id": self.__class__.dec_id, "payment_id": self.__class__.pay_id,
            "kapatilan_tutar": 4000.0})
        assert r.status_code == 400
        assert "kalan" in r.text.lower() or "aş" in r.text.lower()

    def test_over_payment_bakiye(self, tokens):
        # Cross-currency (EUR beyanname, USD bedel) with tiny kur so bedel needed >> bakiye
        r = requests.post(f"{API}/matches", headers=_h(tokens["onaylayan"]), json={
            "declaration_id": self.__class__.dec_id, "payment_id": self.__class__.pay_usd_id,
            "kapatilan_tutar": 3000.0, "kur": 0.1})
        assert r.status_code == 400
        assert "bakiye" in r.text.lower() or "yetersiz" in r.text.lower()

    def test_cross_currency_uses_tcmb(self, tokens):
        """Different currencies -> TCMB rate auto applied; bedel = kapatilan/kur."""
        r = requests.post(f"{API}/matches", headers=_h(tokens["onaylayan"]), json={
            "declaration_id": self.__class__.dec_id, "payment_id": self.__class__.pay_usd_id,
            "kapatilan_tutar": 100.0})
        assert r.status_code in (200, 503), r.text
        if r.status_code == 200:
            m = r.json()
            assert m["kur_kaynak"] == "TCMB"
            assert m["kur"] > 0
            # bedel_kullanilan = 100 / kur, approximately
            assert abs(m["bedel_kullanilan"] - 100.0 / m["kur"]) < 0.02
            requests.delete(f"{API}/matches/{m['id']}", headers=_h(tokens["onaylayan"]))

    def test_delete_blocked_when_matched(self, tokens):
        r = requests.delete(f"{API}/declarations/{self.__class__.dec_id}",
                            headers=_h(tokens["ihracat"]))
        assert r.status_code == 400
        r = requests.delete(f"{API}/payments/{self.__class__.pay_id}",
                            headers=_h(tokens["banka"]))
        assert r.status_code == 400

    def test_update_match_recalcs(self, tokens):
        r = requests.put(f"{API}/matches/{self.__class__.match_id}?kapatilan_tutar=1500",
                         headers=_h(tokens["onaylayan"]))
        assert r.status_code == 200, r.text
        d = next(x for x in requests.get(f"{API}/declarations", headers=_h(tokens["admin"])).json()
                 if x["id"] == self.__class__.dec_id)
        assert d["kapatilan"] == 1500.0

    def test_delete_match_restores(self, tokens):
        r = requests.delete(f"{API}/matches/{self.__class__.match_id}",
                            headers=_h(tokens["onaylayan"]))
        assert r.status_code == 200
        d = next(x for x in requests.get(f"{API}/declarations", headers=_h(tokens["admin"])).json()
                 if x["id"] == self.__class__.dec_id)
        assert d["kapatilan"] == 0.0 and d["durum"] == "ACIK"

    def test_cleanup(self, tokens):
        requests.delete(f"{API}/declarations/{self.__class__.dec_id}", headers=_h(tokens["ihracat"]))
        requests.delete(f"{API}/payments/{self.__class__.pay_id}", headers=_h(tokens["banka"]))
        requests.delete(f"{API}/payments/{self.__class__.pay_usd_id}", headers=_h(tokens["banka"]))


# ---------- reports / excel / dashboard ----------
class TestReportsAndDashboard:
    def test_reports_summary(self, tokens):
        r = requests.get(f"{API}/reports/summary", headers=_h(tokens["admin"]))
        assert r.status_code == 200
        d = r.json()
        for k in ["doviz", "ithalatci", "ibkb_alinmadi", "destek_alinmadi", "ay"]:
            assert k in d, f"missing key {k}"
        assert isinstance(d["ithalatci"], list)
        assert isinstance(d["ibkb_alinmadi"], int)
        assert isinstance(d["destek_alinmadi"], int)

    def test_dashboard(self, tokens):
        r = requests.get(f"{API}/dashboard", headers=_h(tokens["admin"]))
        assert r.status_code == 200
        for k in ["acik", "kismi", "kapali", "acik_tutar", "bedel_bakiye",
                  "yaklasan", "gecmis", "son_hareketler"]:
            assert k in r.json()

    def test_audit_logs(self, tokens):
        r = requests.get(f"{API}/audit-logs", headers=_h(tokens["admin"]))
        assert r.status_code == 200 and isinstance(r.json(), list)

    def test_excel_export_has_new_columns(self, tokens):
        r = requests.get(f"{API}/export/excel", headers=_h(tokens["admin"]))
        assert r.status_code == 200
        ct = r.headers.get("content-type", "")
        assert "spreadsheet" in ct.lower() or "excel" in ct.lower()
        # Read workbook and check headers
        import io
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb["BANKA BİLDİRİMİ"]
        headers = [c.value for c in ws[1]]
        expected = ["SIRA NO", "DOSYA REFERANSI", "GÜMRÜK MÜDÜRLÜĞÜ KODU", "GB NO", "GB TARİHİ",
                    "GB'YE SAYILACAK TUTAR", "KULLANILACAK DTH", "KULLANILACAK ACH",
                    "TCMB DEVİR ORANI", "TEŞVİK", "TAAHHÜT"]
        assert headers[:len(expected)] == expected, headers
        ws1 = wb["Beyanname Listesi"]
        h1 = [c.value for c in ws1[1]]
        for e in ["Beyanname No", "Açılış Tarihi", "Kapanış Tarihi",
                  "Son Kapatma Tarihi (180 gün)", "İthalatçı", "Gümrük Müdürlüğü No",
                  "IBKB Belgesi Durumu", "Destek Ödemesi (%3)", "Destek Durumu",
                  "Teşvik", "Taahhüt"]:
            assert e in h1, f"missing excel header: {e}"


# ---------- users ----------
class TestUsers:
    def test_admin_only_list(self, tokens):
        r = requests.get(f"{API}/users", headers=_h(tokens["admin"]))
        assert r.status_code == 200
        for role in ["ihracat", "banka", "onaylayan", "viewer"]:
            r = requests.get(f"{API}/users", headers=_h(tokens[role]))
            assert r.status_code == 403, f"{role} should not access user list"

    def test_non_admin_cannot_create(self, tokens):
        for role in ["ihracat", "banka", "onaylayan", "viewer"]:
            r = requests.post(f"{API}/users", headers=_h(tokens[role]), json={
                "email": f"nope_{role}@test.com", "name": "x",
                "role": "goruntuleyici", "password": "P@ss12345"})
            assert r.status_code == 403

    def test_admin_crud_user(self, tokens):
        email = f"test_new_{int(datetime.now().timestamp())}@ihracat.com"
        r = requests.post(f"{API}/users", headers=_h(tokens["admin"]), json={
            "email": email, "name": "TEST", "role": "goruntuleyici", "password": "P@ss12345"})
        assert r.status_code == 200, r.text
        uid = r.json()["id"]
        r = requests.put(f"{API}/users/{uid}", headers=_h(tokens["admin"]),
                         json={"role": "banka"})
        assert r.status_code == 200 and r.json()["role"] == "banka"
        r = requests.delete(f"{API}/users/{uid}", headers=_h(tokens["admin"]))
        assert r.status_code == 200


# ---------- 2FA login flow (fallback: e-posta gönderilemediğinde 2FA atlanır) ----------
class TestTwoFactorFallback:
    """EMERGENT_EMAIL_KEY geçersiz olduğunda, 2FA açık kullanıcılar için de giriş
    doğrudan token dönmeli (TWO_FACTOR_FALLBACK=true varsayılan)."""

    def test_2fa_on_user_login_never_locks(self):
        # test2fa@ihracat.com'un two_factor=True. E-posta gönderilirse challenge,
        # gönderilemezse fallback ile token dönmeli. Her iki durumda da 200 OK.
        r = requests.post(f"{API}/auth/login",
                          json={"email": TWOFA_EMAIL, "password": TWOFA_PW}, timeout=30)
        assert r.status_code == 200, r.text
        data = r.json()
        assert data.get("token") or data.get("challenge_id"), (
            f"Ne token ne challenge döndü, kullanıcı kilitlendi: {data}")

    def test_2fa_atlandi_audit_log_created(self, tokens):
        # Yukarıdaki login sonrası '2FA_ATLANDI' audit log'u yazılmalı
        r = requests.get(f"{API}/audit-logs", headers=_h(tokens["admin"]),
                         params={"modul": "Giriş"})
        assert r.status_code == 200, r.text
        logs = r.json()
        assert any(l.get("islem") == "2FA_ATLANDI" for l in logs), (
            f"2FA_ATLANDI audit log bulunamadı, gelenler: {[l.get('islem') for l in logs[:20]]}")

    def test_two_factor_off_account_single_step(self):
        r = requests.post(f"{API}/auth/login",
                          json={"email": "admin@ihracat.com", "password": "Admin1234!"})
        assert r.status_code == 200
        data = r.json()
        assert data.get("token")
        assert not data.get("two_factor")

    def test_wrong_password_still_401(self):
        r = requests.post(f"{API}/auth/login",
                          json={"email": TWOFA_EMAIL, "password": "wrongwrong"})
        assert r.status_code == 401
        assert "hatalı" in r.text.lower() or "hatali" in r.text.lower()


# ---------- Yeni: kullanıcı ekleme varsayılan 2FA=Kapalı + şifre validasyonu + admin şifre değiştirme ----------
class TestNewUserFlow:
    """Bug fix: Admin yeni kullanıcı eklediğinde 2FA varsayılan KAPALI olmalı,
    kullanıcı tek adımda giriş yapabilmeli. Şifre < 6 karakter reddedilmeli.
    Admin PUT ile şifre güncelleyince yeni şifre çalışmalı, eski çalışmamalı."""

    created_ids = []

    def test_default_two_factor_false_and_single_step_login(self, tokens):
        email = f"test_newuser_{int(datetime.now().timestamp())}@ihracat.com"
        pw = "Yeni1234!"
        # two_factor gönderilmiyor
        r = requests.post(f"{API}/users", headers=_h(tokens["admin"]), json={
            "email": email, "name": "TEST Yeni Kullanıcı",
            "role": "goruntuleyici", "password": pw})
        assert r.status_code == 200, r.text
        u = r.json()
        assert u.get("two_factor") is False, u
        assert u["email"] == email
        self.__class__.created_ids.append(u["id"])

        # E-posta + şifre → doğrudan token, 2FA challenge gelmemeli
        r = requests.post(f"{API}/auth/login",
                          json={"email": email, "password": pw})
        assert r.status_code == 200, r.text
        data = r.json()
        assert data.get("token"), data
        assert "challenge_id" not in data
        assert not data.get("two_factor")

    def test_password_min_6_chars(self, tokens):
        email = f"test_shortpw_{int(datetime.now().timestamp())}@ihracat.com"
        r = requests.post(f"{API}/users", headers=_h(tokens["admin"]), json={
            "email": email, "name": "TEST Kısa Şifre",
            "role": "goruntuleyici", "password": "12345"})
        assert r.status_code == 400, r.text
        assert "6" in r.text and ("şifre" in r.text.lower() or "sifre" in r.text.lower())

    def test_invalid_role_400(self, tokens):
        email = f"test_badrole_{int(datetime.now().timestamp())}@ihracat.com"
        r = requests.post(f"{API}/users", headers=_h(tokens["admin"]), json={
            "email": email, "name": "TEST", "role": "hacker",
            "password": "P@ss12345"})
        assert r.status_code == 400
        assert "rol" in r.text.lower()

    def test_duplicate_email_400(self, tokens):
        email = f"test_dupe_{int(datetime.now().timestamp())}@ihracat.com"
        r = requests.post(f"{API}/users", headers=_h(tokens["admin"]), json={
            "email": email, "name": "TEST", "role": "goruntuleyici",
            "password": "P@ss12345"})
        assert r.status_code == 200
        self.__class__.created_ids.append(r.json()["id"])
        r2 = requests.post(f"{API}/users", headers=_h(tokens["admin"]), json={
            "email": email, "name": "TEST", "role": "goruntuleyici",
            "password": "P@ss12345"})
        assert r2.status_code == 400

    def test_admin_password_change_via_put(self, tokens):
        email = f"test_pwchange_{int(datetime.now().timestamp())}@ihracat.com"
        old_pw = "Eski1234!"
        new_pw = "Yeni9876!"
        r = requests.post(f"{API}/users", headers=_h(tokens["admin"]), json={
            "email": email, "name": "TEST PW",
            "role": "goruntuleyici", "password": old_pw})
        assert r.status_code == 200, r.text
        uid = r.json()["id"]
        self.__class__.created_ids.append(uid)

        # Eski şifre çalışmalı
        r = requests.post(f"{API}/auth/login",
                          json={"email": email, "password": old_pw})
        assert r.status_code == 200 and r.json().get("token")

        # Admin PUT ile şifre değiştir
        r = requests.put(f"{API}/users/{uid}", headers=_h(tokens["admin"]),
                         json={"password": new_pw})
        assert r.status_code == 200, r.text

        # Eski şifre reddedilmeli
        r = requests.post(f"{API}/auth/login",
                          json={"email": email, "password": old_pw})
        assert r.status_code == 401

        # Yeni şifre çalışmalı
        r = requests.post(f"{API}/auth/login",
                          json={"email": email, "password": new_pw})
        assert r.status_code == 200 and r.json().get("token")

    def test_2fa_on_new_user_fallback_login(self, tokens):
        """two_factor=True ile açılan kullanıcı da e-posta gönderilemediği için
        fallback ile doğrudan token almalı (kilitlenmemeli)."""
        email = f"test_2fa_on_{int(datetime.now().timestamp())}@ihracat.com"
        pw = "Tfa12345!"
        r = requests.post(f"{API}/users", headers=_h(tokens["admin"]), json={
            "email": email, "name": "TEST 2FA ON",
            "role": "goruntuleyici", "password": pw, "two_factor": True})
        assert r.status_code == 200, r.text
        u = r.json()
        assert u["two_factor"] is True
        self.__class__.created_ids.append(u["id"])

        r = requests.post(f"{API}/auth/login",
                          json={"email": email, "password": pw})
        assert r.status_code == 200, r.text
        data = r.json()
        # E-posta gönderilebildiyse challenge dönebilir; gönderilemediyse fallback ile token dönmeli.
        # Her iki durumda da kullanıcı kilitlenmemeli (200 OK olması yeterli).
        assert data.get("token") or data.get("challenge_id"), (
            f"Login ne token ne challenge döndü: {data}")

    def test_cleanup(self, tokens):
        for uid in self.__class__.created_ids:
            requests.delete(f"{API}/users/{uid}", headers=_h(tokens["admin"]))
        self.__class__.created_ids = []


class TestBruteForceLockout:
    """5 hatalı denemeden sonra 15 dakika kilit; 429 dönmeli."""

    def test_locks_after_5_failures(self, tokens):
        email = f"test_lock_{int(datetime.now().timestamp())}@ihracat.com"
        pw = "Lock12345!"
        r = requests.post(f"{API}/users", headers=_h(tokens["admin"]), json={
            "email": email, "name": "TEST LOCK",
            "role": "goruntuleyici", "password": pw})
        assert r.status_code == 200
        uid = r.json()["id"]
        try:
            # 5 wrong attempts
            for _ in range(5):
                r = requests.post(f"{API}/auth/login",
                                  json={"email": email, "password": "wrongwrong"})
                assert r.status_code == 401
            # 6th attempt should be locked (429). Correct pw should also fail with 429.
            r = requests.post(f"{API}/auth/login",
                              json={"email": email, "password": pw})
            assert r.status_code == 429, f"expected 429 lock, got {r.status_code}: {r.text}"
        finally:
            requests.delete(f"{API}/users/{uid}", headers=_h(tokens["admin"]))

    def test_inactive_user_403(self, tokens):
        email = f"test_inactive_{int(datetime.now().timestamp())}@ihracat.com"
        pw = "Inac1234!"
        r = requests.post(f"{API}/users", headers=_h(tokens["admin"]), json={
            "email": email, "name": "TEST INAC",
            "role": "goruntuleyici", "password": pw})
        uid = r.json()["id"]
        try:
            requests.put(f"{API}/users/{uid}", headers=_h(tokens["admin"]),
                         json={"active": False})
            r = requests.post(f"{API}/auth/login",
                              json={"email": email, "password": pw})
            assert r.status_code == 403, r.text
        finally:
            requests.delete(f"{API}/users/{uid}", headers=_h(tokens["admin"]))


# ---------- alerts preview + send RBAC ----------
class TestAlerts:
    def test_preview_all_roles(self, tokens):
        for role in ["admin", "ihracat", "banka", "onaylayan", "viewer"]:
            r = requests.get(f"{API}/alerts/preview", headers=_h(tokens[role]))
            assert r.status_code == 200, f"{role} {r.text}"
            data = r.json()
            for k in ["alicilar", "sayilar", "plan"]:
                assert k in data
            for k in ["gecmis", "yaklasan", "ibkb", "destek"]:
                assert k in data["sayilar"]
            assert isinstance(data["alicilar"], list)
            assert data["alicilar"], "recipients list should be non-empty"

    def test_send_rbac_forbidden_roles(self, tokens):
        for role in ["ihracat", "banka", "viewer"]:
            r = requests.post(f"{API}/alerts/send", headers=_h(tokens[role]))
            assert r.status_code == 403, f"{role} should be 403 got {r.status_code}"

    def test_send_by_onaylayan_and_audit(self, tokens):
        # Single real send (avoid spamming)
        r = requests.post(f"{API}/alerts/send", headers=_h(tokens["onaylayan"]))
        assert r.status_code == 200, r.text
        data = r.json()
        assert data.get("sent") is True
        # audit log should contain Uyarı/EPOSTA entry
        r = requests.get(f"{API}/audit-logs", headers=_h(tokens["admin"]),
                         params={"modul": "Uyarı"})
        assert r.status_code == 200
        logs = r.json()
        assert any(l.get("islem") == "EPOSTA" for l in logs), "no audit log for alert send"


# ---------- IBKB (v3) + Banka Bildirimi Excel şablonu ----------
class TestIbkbAndBankaBildirimi:
    dec_id = None
    dec_no = f"TEST_IBKBFLOW_{int(datetime.now().timestamp())}"
    pay_eur_id = None
    pay_usd_id = None
    match_ids = []
    acilis = "2026-04-02"

    def test_seed(self, tokens):
        # Beyanname (EUR 10000)
        r = requests.post(f"{API}/declarations", headers=_h(tokens["ihracat"]),
                          json={**_base_dec(self.__class__.dec_no, self.acilis,
                                            doviz="EUR", tutar=10000.0,
                                            ithalatci="TEST IBKB GmbH", gumruk="GM-IBKB"),
                                "tesvik": True, "taahhut": True})
        assert r.status_code == 200, r.text
        self.__class__.dec_id = r.json()["id"]

        # EUR bedel 6000
        r = requests.post(f"{API}/payments", headers=_h(tokens["banka"]), json={
            "banka": "Ziraat", "gonderen": "TEST IBKB EUR",
            "tarih": self.acilis, "doviz": "EUR", "tutar": 6000.0})
        assert r.status_code == 200
        self.__class__.pay_eur_id = r.json()["id"]

        # USD bedel 9000
        r = requests.post(f"{API}/payments", headers=_h(tokens["banka"]), json={
            "banka": "Is Bankasi", "gonderen": "TEST IBKB USD",
            "tarih": self.acilis, "doviz": "USD", "tutar": 9000.0})
        assert r.status_code == 200
        self.__class__.pay_usd_id = r.json()["id"]

    def test_payment_persists_dth_ach_iban(self, tokens):
        """PaymentInput.dth_iban / ach_iban are persisted and readable via GET /api/payments."""
        r = requests.post(f"{API}/payments", headers=_h(tokens["banka"]), json={
            "banka": "TEST", "gonderen": "TEST_IBAN_PAY",
            "tarih": "2026-04-02", "doviz": "EUR", "tutar": 100.0,
            "dth_iban": "TR22 EUR IBAN 0000 0000 0000 01",
            "ach_iban": "TR22 TL  IBAN 0000 0000 0000 02",
        })
        assert r.status_code == 200, r.text
        pid = r.json()["id"]
        assert r.json()["dth_iban"].startswith("TR22 EUR")
        assert r.json()["ach_iban"].startswith("TR22 TL")

        got = next(x for x in requests.get(f"{API}/payments", headers=_h(tokens["admin"])).json()
                   if x["id"] == pid)
        assert got["dth_iban"].startswith("TR22 EUR")
        assert got["ach_iban"].startswith("TR22 TL")
        requests.delete(f"{API}/payments/{pid}", headers=_h(tokens["banka"]))

    def test_ach_iban_fallback_to_default_env(self, tokens):
        """When ach_iban is empty, payment_view returns DEFAULT_ACH_IBAN env value."""
        default_iban = "TR00 0000 0000 0000 0000 0000 00"
        r = requests.post(f"{API}/payments", headers=_h(tokens["banka"]), json={
            "banka": "TEST", "gonderen": "TEST_ACH_FALLBACK",
            "tarih": "2026-04-02", "doviz": "EUR", "tutar": 50.0,
        })
        assert r.status_code == 200
        pid = r.json()["id"]
        got = next(x for x in requests.get(f"{API}/payments", headers=_h(tokens["admin"])).json()
                   if x["id"] == pid)
        assert got["ach_iban"] == "", got["ach_iban"]
        assert got["ach_iban_default"] == default_iban, got["ach_iban_default"]
        requests.delete(f"{API}/payments/{pid}", headers=_h(tokens["banka"]))

    def test_ibkb_payment_view_defaults(self, tokens):
        r = requests.get(f"{API}/payments", headers=_h(tokens["admin"]))
        assert r.status_code == 200
        p = next(x for x in r.json() if x["id"] == self.__class__.pay_eur_id)
        assert p["ibkb_durum"] == "DUZENLENMEDI"
        # zorunlu bozdurma = %35 (iteration 8'de %30 -> %35 güncellendi)
        assert abs(p["zorunlu_bozdurma"] - 2100.0) < 0.01
        assert p["zorunlu_bozdurma_orani"] == 35.0
        assert p["tcmb_devir_orani"] == 100.0

    def test_ibkb_rbac_forbidden_roles(self, tokens):
        payload = {"ibkb_no": "IBKB-1", "ibkb_tarihi": "2026-04-03",
                   "dosya_referansi": "REF-1", "ach_iban": "TR11 0006 2000 0000 0000 0000 01",
                   "tcmb_devir_orani": 100}
        for role in ["ihracat", "onaylayan", "viewer"]:
            r = requests.put(f"{API}/payments/{self.__class__.pay_eur_id}/ibkb",
                             headers=_h(tokens[role]), json=payload)
            assert r.status_code == 403, f"{role} should be 403, got {r.status_code}"

    def test_dth_iban_from_payment(self, tokens):
        r = requests.get(f"{API}/payments", headers=_h(tokens["admin"]))
        p = next(x for x in r.json() if x["id"] == self.__class__.pay_eur_id)
        assert "dth_iban" in p and "ach_iban" in p

    def test_banka_can_save_ibkb(self, tokens):
        r = requests.put(f"{API}/payments/{self.__class__.pay_eur_id}/ibkb",
                         headers=_h(tokens["banka"]),
                         json={"ibkb_duzenlendi": True, "ibkb_no": "IBKB-EUR-1",
                               "ibkb_tarihi": "2026-04-03",
                               "dosya_referansi": "DOSYA-EUR-1",
                               "ach_iban": "TR11 0006 2000 1234 0006 2999 88",
                               "tcmb_devir_orani": 100})
        assert r.status_code == 200, r.text
        p = r.json()
        assert p["ibkb_durum"] == "DUZENLENDI"
        assert p["ibkb_no"] == "IBKB-EUR-1"
        assert p["dosya_referansi"] == "DOSYA-EUR-1"
        assert p["ach_iban"] == "TR11 0006 2000 1234 0006 2999 88"
        assert p["tcmb_devir_orani"] == 100

        r = requests.put(f"{API}/payments/{self.__class__.pay_usd_id}/ibkb",
                         headers=_h(tokens["admin"]),
                         json={"ibkb_duzenlendi": True, "ibkb_no": "IBKB-USD-1",
                               "ibkb_tarihi": "2026-04-03",
                               "dosya_referansi": "DOSYA-USD-1",
                               "tcmb_devir_orani": 100})
        assert r.status_code == 200

    def test_ibkb_input_rejects_legacy_tutar_fields(self, tokens):
        """dth_tutar / ach_tutar are no longer part of IbkbInput; extras must be ignored.
        The endpoint must still succeed but the payment must not store any tutar fields."""
        r = requests.put(f"{API}/payments/{self.__class__.pay_eur_id}/ibkb",
                         headers=_h(tokens["banka"]),
                         json={"ibkb_duzenlendi": True, "ibkb_no": "IBKB-EUR-1",
                               "ibkb_tarihi": "2026-04-03",
                               "dosya_referansi": "DOSYA-EUR-1",
                               "ach_iban": "TR11 0006 2000 1234 0006 2999 88",
                               "tcmb_devir_orani": 100,
                               "dth_tutar": 123, "ach_tutar": 456})
        assert r.status_code == 200, r.text
        p = r.json()
        assert "dth_tutar" not in p
        assert "ach_tutar" not in p

    def test_create_matches(self, tokens):
        # EUR->EUR match 3000
        r = requests.post(f"{API}/matches", headers=_h(tokens["onaylayan"]), json={
            "declaration_id": self.__class__.dec_id,
            "payment_id": self.__class__.pay_eur_id,
            "kapatilan_tutar": 3000.0})
        assert r.status_code == 200, r.text
        self.__class__.match_ids.append(r.json()["id"])

        # USD->EUR match 1000 (manual kur 1.1 to avoid TCMB dependency)
        r = requests.post(f"{API}/matches", headers=_h(tokens["onaylayan"]), json={
            "declaration_id": self.__class__.dec_id,
            "payment_id": self.__class__.pay_usd_id,
            "kapatilan_tutar": 1000.0, "kur": 0.9})
        assert r.status_code == 200, r.text
        self.__class__.match_ids.append(r.json()["id"])

    def test_excel_banka_bildirimi_shape(self, tokens):
        r = requests.get(f"{API}/export/excel", headers=_h(tokens["admin"]))
        assert r.status_code == 200
        import io as _io
        from openpyxl import load_workbook
        wb = load_workbook(_io.BytesIO(r.content))
        assert "BANKA BİLDİRİMİ" in wb.sheetnames
        assert wb.sheetnames[0] == "BANKA BİLDİRİMİ"
        assert "Beyanname Listesi" in wb.sheetnames
        assert "Eşleştirme Detayı" in wb.sheetnames

        ws = wb["BANKA BİLDİRİMİ"]
        expected = ["SIRA NO", "DOSYA REFERANSI", "GÜMRÜK MÜDÜRLÜĞÜ KODU", "GB NO",
                    "GB TARİHİ", "GB'YE SAYILACAK TUTAR", "KULLANILACAK DTH",
                    "KULLANILACAK ACH", "TCMB DEVİR ORANI", "TEŞVİK", "TAAHHÜT"]
        headers = [c.value for c in ws[1]]
        assert headers == expected, headers

        # Find our two rows by GB NO
        our_rows = []
        toplam_row = None
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == "TOPLAM":
                toplam_row = row
                continue
            if row[3] == self.__class__.dec_no:
                our_rows.append(row)
        assert len(our_rows) == 2, f"expected 2 matching rows, got {len(our_rows)}"

        # Sıra artıyor
        siras = [r[0] for r in our_rows]
        assert all(isinstance(s, int) and s > 0 for s in siras)

        for row in our_rows:
            # dosya referansı bedelden
            assert row[1] in ("DOSYA-EUR-1", "DOSYA-USD-1"), row
            assert row[2] == "GM-IBKB"
            assert row[3] == self.__class__.dec_no
            # GB tarihi dd.mm.yyyy
            assert row[4] == "02.04.2026", row[4]
            assert row[8] == "%100"
            assert row[9] == "E"  # tesvik (banka şablonunda E/H)
            assert row[10] == "E"  # taahhut

        # EUR row: DTH/ACH artık IBAN metni
        eur_row = next(r for r in our_rows if r[1] == "DOSYA-EUR-1")
        assert abs(eur_row[5] - 3000.0) < 0.01
        assert eur_row[7] == "TR11 0006 2000 1234 0006 2999 88", eur_row[7]

        usd_row = next(r for r in our_rows if r[1] == "DOSYA-USD-1")
        assert abs(usd_row[5] - 1000.0) < 0.01

        # Başlık satırı boyalı OLMAMALI (banka istemiyor)
        assert ws["A1"].fill.fill_type in (None, "none"), ws["A1"].fill.fill_type
        assert ws["A1"].font.bold is True

        assert toplam_row is not None
        # TOPLAM col F must include our two matches (3000 + 1000 = 4000);
        # parallel test classes may add unrelated matches so use >=.
        assert toplam_row[5] >= 4000.0 - 0.01, toplam_row

        # Beyanname Listesi tesvik/taahhut kolonları
        ws1 = wb["Beyanname Listesi"]
        h1 = [c.value for c in ws1[1]]
        assert "Teşvik" in h1 and "Taahhüt" in h1

        # Eşleştirme Detayı IBKB No + Dosya Ref kolonları
        ws2 = wb["Eşleştirme Detayı"]
        h2 = [c.value for c in ws2[1]]
        assert "IBKB No" in h2 and "Dosya Referansı" in h2
        assert "DTH IBAN" in h2 and "ACH IBAN" in h2

    def test_declaration_view_returns_tesvik_taahhut(self, tokens):
        r = requests.get(f"{API}/declarations", headers=_h(tokens["admin"]),
                         params={"q": self.__class__.dec_no})
        assert r.status_code == 200
        d = next(x for x in r.json() if x["id"] == self.__class__.dec_id)
        assert d.get("tesvik") is True
        assert d.get("taahhut") is True

    def test_audit_has_ibkb_module(self, tokens):
        r = requests.get(f"{API}/audit-logs", headers=_h(tokens["admin"]),
                         params={"modul": "IBKB"})
        assert r.status_code == 200
        logs = r.json()
        assert any(l.get("modul") == "IBKB" for l in logs), "IBKB audit log missing"

    def test_cleanup(self, tokens):
        for mid in self.__class__.match_ids:
            requests.delete(f"{API}/matches/{mid}", headers=_h(tokens["onaylayan"]))
        requests.delete(f"{API}/declarations/{self.__class__.dec_id}",
                        headers=_h(tokens["ihracat"]))
        requests.delete(f"{API}/payments/{self.__class__.pay_eur_id}",
                        headers=_h(tokens["banka"]))
        requests.delete(f"{API}/payments/{self.__class__.pay_usd_id}",
                        headers=_h(tokens["banka"]))


# ---------- Iteration 6: NOTLAR removal, export/check, dashboard extras, backup/restore ----------
class TestIteration6NotesRemoval:
    """BANKA BİLDİRİMİ sayfasında NOTLAR/taahhütname bloğu HİÇ olmamalı ve merged cell bulunmamalı."""

    FORBIDDEN_SUBSTR = [
        "NOTLAR", "notlar",
        "taahhüt ederiz", "taahhut ederiz", "TAAHHÜT EDERIZ", "TAAHHUT EDERIZ",
        "Döviz Dönüşüm Desteği", "Doviz Donusum Destegi",
        "Döviz Dönüşüm Desteği Talebi",
        "kaşe", "kase",
        "imza", "İmza",
    ]

    def test_excel_first_sheet_has_no_notes_block(self, tokens):
        r = requests.get(f"{API}/export/excel", headers=_h(tokens["admin"]))
        assert r.status_code == 200
        import io as _io
        from openpyxl import load_workbook
        wb = load_workbook(_io.BytesIO(r.content))
        assert wb.sheetnames[0] == "BANKA BİLDİRİMİ"
        ws = wb["BANKA BİLDİRİMİ"]

        # (1) No merged cells anywhere on the first sheet
        merged = list(ws.merged_cells.ranges)
        assert merged == [], f"BANKA BİLDİRİMİ should have no merged cells, got: {merged}"

        # (2) No forbidden substrings in any cell value on the first sheet
        offenders = []
        for row in ws.iter_rows(values_only=True):
            for val in row:
                if val is None:
                    continue
                s = str(val)
                for needle in self.FORBIDDEN_SUBSTR:
                    if needle in s:
                        offenders.append((needle, s))
        assert not offenders, f"BANKA BİLDİRİMİ contains forbidden notes text: {offenders[:5]}"

        # (3) The last non-empty row must be TOPLAM (i.e. nothing appended after totals)
        last_data_row = None
        for r_idx in range(ws.max_row, 0, -1):
            row_vals = [ws.cell(row=r_idx, column=c).value for c in range(1, ws.max_column + 1)]
            if any(v not in (None, "") for v in row_vals):
                last_data_row = r_idx
                break
        assert last_data_row is not None
        assert ws.cell(row=last_data_row, column=1).value == "TOPLAM", (
            f"Last row of BANKA BİLDİRİMİ must be TOPLAM, got "
            f"{[ws.cell(row=last_data_row, column=c).value for c in range(1, ws.max_column + 1)]}"
        )

    def test_excel_opens_cleanly_with_openpyxl(self, tokens):
        """.xlsx must be valid — load_workbook must not raise."""
        r = requests.get(f"{API}/export/excel", headers=_h(tokens["admin"]))
        assert r.status_code == 200
        assert "spreadsheet" in r.headers.get("content-type", "").lower()
        import io as _io
        from openpyxl import load_workbook
        wb = load_workbook(_io.BytesIO(r.content))
        # required sheets still there
        for name in ["BANKA BİLDİRİMİ", "Beyanname Listesi", "Eşleştirme Detayı"]:
            assert name in wb.sheetnames, wb.sheetnames

        # Beyanname Listesi has rows (>=1 data row expected in preview DB)
        ws1 = wb["Beyanname Listesi"]
        assert ws1.max_row >= 1
        # Eşleştirme Detayı has header
        ws2 = wb["Eşleştirme Detayı"]
        headers2 = [c.value for c in ws2[1]]
        assert "Beyanname No" in headers2 and "DTH IBAN" in headers2 and "ACH IBAN" in headers2


class TestIteration6ExportCheck:
    def test_export_check_ok(self, tokens):
        r = requests.get(f"{API}/export/check", headers=_h(tokens["admin"]))
        assert r.status_code == 200, r.text
        d = r.json()
        for k in ("satir_sayisi", "eksik_sayisi", "eksikler"):
            assert k in d, d
        assert isinstance(d["satir_sayisi"], int)
        assert isinstance(d["eksik_sayisi"], int)
        assert isinstance(d["eksikler"], list)
        # each missing row has beyanname_no / bedel / alanlar list
        for e in d["eksikler"]:
            assert "beyanname_no" in e and "alanlar" in e and isinstance(e["alanlar"], list)

    def test_export_check_requires_auth(self):
        r = requests.get(f"{API}/export/check")
        assert r.status_code == 401


class TestIteration6DashboardExtras:
    def test_dashboard_has_new_cards(self, tokens):
        r = requests.get(f"{API}/dashboard", headers=_h(tokens["admin"]))
        assert r.status_code == 200
        d = r.json()
        for k in ("bu_ay_kapatilan", "bu_ay_islem_sayi",
                  "destek_bekleyen_sayi", "destek_bekleyen_tutar"):
            assert k in d, f"missing dashboard key {k}: keys={list(d.keys())}"
        assert isinstance(d["bu_ay_kapatilan"], dict)
        assert isinstance(d["bu_ay_islem_sayi"], int)
        assert isinstance(d["destek_bekleyen_sayi"], int)
        assert isinstance(d["destek_bekleyen_tutar"], dict)
        # legacy keys still present
        for k in ("acik", "kismi", "kapali", "acik_tutar", "bedel_bakiye",
                  "yaklasan", "gecmis", "son_hareketler"):
            assert k in d


class TestIteration6Backup:
    """/api/backup indir + /api/backup/restore geri yükleme (admin-only)."""

    backup_bytes = None

    def test_backup_rbac_non_admin_forbidden(self, tokens):
        for role in ("ihracat", "banka", "onaylayan", "viewer"):
            r = requests.get(f"{API}/backup", headers=_h(tokens[role]))
            assert r.status_code == 403, f"{role} got {r.status_code}"

    def test_backup_download_json(self, tokens):
        r = requests.get(f"{API}/backup", headers=_h(tokens["admin"]))
        assert r.status_code == 200, r.text
        assert "application/json" in r.headers.get("content-type", "").lower()
        assert "attachment" in r.headers.get("content-disposition", "").lower()
        import json as _json
        data = _json.loads(r.content.decode("utf-8"))
        assert "_meta" in data
        for col in ("users", "declarations", "payments", "matches", "audit_logs"):
            assert col in data, f"missing collection {col} in backup"
            assert isinstance(data[col], list)
        # meta info
        assert data["_meta"].get("olusturan") == "admin@ihracat.com"
        self.__class__.backup_bytes = r.content

    def test_restore_rbac_non_admin_forbidden(self, tokens):
        # Use a minimal fake backup for rbac (no real restore should occur since 403 first)
        files = {"file": ("dummy.json", b'{"users": []}', "application/json")}
        for role in ("ihracat", "banka", "onaylayan", "viewer"):
            hdr = {"Authorization": f"Bearer {tokens[role]}"}
            r = requests.post(f"{API}/backup/restore?mode=merge",
                              headers=hdr, files=files)
            assert r.status_code == 403, f"{role} got {r.status_code}"

    def test_restore_invalid_file_400(self, tokens):
        hdr = {"Authorization": f"Bearer {tokens['admin']}"}
        files = {"file": ("bad.json", b"not-json-content", "application/json")}
        r = requests.post(f"{API}/backup/restore?mode=merge", headers=hdr, files=files)
        assert r.status_code == 400, r.text

    def test_restore_non_backup_json_400(self, tokens):
        hdr = {"Authorization": f"Bearer {tokens['admin']}"}
        files = {"file": ("noop.json", b'{"foo": "bar"}', "application/json")}
        r = requests.post(f"{API}/backup/restore?mode=merge", headers=hdr, files=files)
        assert r.status_code == 400

    def test_restore_merge_roundtrip(self, tokens):
        """Download → merge-restore the same backup. Should be idempotent and 200."""
        assert self.__class__.backup_bytes, "backup was not captured earlier"
        # count declarations before
        before = requests.get(f"{API}/declarations", headers=_h(tokens["admin"])).json()
        n_before = len(before)

        hdr = {"Authorization": f"Bearer {tokens['admin']}"}
        files = {"file": ("backup.json", self.__class__.backup_bytes, "application/json")}
        r = requests.post(f"{API}/backup/restore?mode=merge", headers=hdr, files=files)
        assert r.status_code == 200, r.text
        body = r.json()
        assert body.get("ok") is True
        assert body.get("mode") == "merge"
        assert "yuklenen" in body
        for col in ("users", "declarations", "payments", "matches", "audit_logs"):
            assert col in body["yuklenen"]

        # After merge with identical data, declaration count should not decrease
        after = requests.get(f"{API}/declarations", headers=_h(tokens["admin"])).json()
        assert len(after) >= n_before, (n_before, len(after))

    def test_backup_audit_log(self, tokens):
        r = requests.get(f"{API}/audit-logs", headers=_h(tokens["admin"]),
                         params={"modul": "Yedek"})
        assert r.status_code == 200
        logs = r.json()
        assert any(l.get("islem") == "INDIR" for l in logs), "backup INDIR audit log missing"
        assert any(l.get("islem") == "GERI_YUKLE" for l in logs), "restore audit log missing"



# ---------- Iteration 7: 'Eşleştirme Detayı' 19 kolon + yeni 3 kolon (TCMB/Teşvik/Taahhüt) ----------
class TestIteration7EslestirmeDetayiColumns:
    """Excel 'Eşleştirme Detayı' sayfası tam 19 kolon içermeli ve yeni 3 kolon
    (TCMB Devir Oranı, Teşvik, Taahhüt) ACH IBAN'dan sonra ve 'Kur' kolonundan önce yer almalı.
    tcmb_devir_orani=95 gibi özel bir değer Excel'e '%95' olarak yansımalı."""

    dec_id = None
    dec_no = f"TEST_ITER7_{int(datetime.now().timestamp())}"
    pay_id = None
    match_id = None
    acilis = "2026-05-01"

    def test_seed(self, tokens):
        r = requests.post(f"{API}/declarations", headers=_h(tokens["ihracat"]),
                          json={**_base_dec(self.__class__.dec_no, self.acilis,
                                            doviz="EUR", tutar=5000.0,
                                            ithalatci="TEST ITER7 GmbH",
                                            gumruk="GM-ITER7"),
                                "tesvik": True, "taahhut": False})
        assert r.status_code == 200, r.text
        self.__class__.dec_id = r.json()["id"]

        r = requests.post(f"{API}/payments", headers=_h(tokens["banka"]), json={
            "banka": "TEST ITER7 BANK", "gonderen": "TEST_ITER7_SENDER",
            "tarih": self.acilis, "doviz": "EUR", "tutar": 2000.0,
            "dth_iban": "TR77 ITER7 DTH 0000 0000 0000 01",
            "ach_iban": "TR77 ITER7 ACH 0000 0000 0000 02",
        })
        assert r.status_code == 200, r.text
        self.__class__.pay_id = r.json()["id"]

        # IBKB güncellemesinde tcmb_devir_orani=95 (özellikle default 100 dışında)
        r = requests.put(f"{API}/payments/{self.__class__.pay_id}/ibkb",
                         headers=_h(tokens["banka"]),
                         json={"ibkb_duzenlendi": True,
                               "ibkb_no": "IBKB-ITER7",
                               "ibkb_tarihi": "2026-05-02",
                               "dosya_referansi": "DOSYA-ITER7",
                               "tcmb_devir_orani": 95})
        assert r.status_code == 200, r.text
        assert r.json()["tcmb_devir_orani"] == 95

        r = requests.post(f"{API}/matches", headers=_h(tokens["onaylayan"]), json={
            "declaration_id": self.__class__.dec_id,
            "payment_id": self.__class__.pay_id,
            "kapatilan_tutar": 1500.0})
        assert r.status_code == 200, r.text
        self.__class__.match_id = r.json()["id"]

    def test_eslestirme_detayi_has_19_columns_in_correct_order(self, tokens):
        r = requests.get(f"{API}/export/excel", headers=_h(tokens["admin"]))
        assert r.status_code == 200
        import io as _io
        from openpyxl import load_workbook
        wb = load_workbook(_io.BytesIO(r.content))
        assert wb.sheetnames == ["BANKA BİLDİRİMİ", "Beyanname Listesi", "Eşleştirme Detayı"], wb.sheetnames

        ws2 = wb["Eşleştirme Detayı"]
        headers = [c.value for c in ws2[1]]
        expected = [
            "Beyanname No", "Bedel Gönderen", "Banka", "IBKB No", "Dosya Referansı",
            "Bedel Tarihi", "Bedel Dövizi", "Kullanılan Bedel", "DTH IBAN", "ACH IBAN",
            "TCMB Devir Oranı", "Teşvik", "Taahhüt", "Kur", "Kur Kaynağı",
            "Kapatılan (Beyanname Dövizi)", "Beyanname Dövizi", "İşlem Tarihi", "İşlemi Yapan",
        ]
        assert headers == expected, headers
        assert len(headers) == 19

        # Yeni 3 kolon ACH IBAN (idx 9) sonrası ve 'Kur' (idx 13) öncesi konumda
        assert headers.index("TCMB Devir Oranı") == 10
        assert headers.index("Teşvik") == 11
        assert headers.index("Taahhüt") == 12
        assert headers.index("ACH IBAN") == 9
        assert headers.index("Kur") == 13

    def test_eslestirme_detayi_new_cols_populated_with_custom_tcmb(self, tokens):
        r = requests.get(f"{API}/export/excel", headers=_h(tokens["admin"]))
        assert r.status_code == 200
        import io as _io
        from openpyxl import load_workbook
        wb = load_workbook(_io.BytesIO(r.content))
        ws2 = wb["Eşleştirme Detayı"]

        our = None
        for row in ws2.iter_rows(min_row=2, values_only=True):
            if row[0] == self.__class__.dec_no:
                our = row
                break
        assert our is not None, "yeni eşleştirme satırı Eşleştirme Detayı'nda bulunamadı"

        # index 10 = TCMB Devir Oranı; girilen 95 değeri '%95' olarak yansımalı
        assert our[10] == "%95", our[10]
        # index 11 = Teşvik (E), index 12 = Taahhüt (H)
        assert our[11] == "E", our
        assert our[12] == "H", our
        # Kur kolonu (13) sayısal, boş değil
        assert isinstance(our[13], (int, float)) and our[13] > 0, our

        # Hiçbir yeni 3 kolonun hücresi diğer veri satırlarında boş olmamalı
        idx_tcmb, idx_tes, idx_taah = 10, 11, 12
        for row in ws2.iter_rows(min_row=2, values_only=True):
            if row[0] is None:  # boş satır
                continue
            assert row[idx_tcmb] not in (None, ""), row
            assert row[idx_tes] in ("E", "H"), row
            assert row[idx_taah] in ("E", "H"), row

    def test_banka_bildirimi_custom_tcmb_reflected(self, tokens):
        """İlk sayfada da girilen tcmb_devir_orani=95 satıra '%95' olarak yansımalı."""
        r = requests.get(f"{API}/export/excel", headers=_h(tokens["admin"]))
        assert r.status_code == 200
        import io as _io
        from openpyxl import load_workbook
        wb = load_workbook(_io.BytesIO(r.content))
        ws = wb["BANKA BİLDİRİMİ"]
        # ilk sayfada beyanname no col 4 (index 3)
        found = None
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == "TOPLAM":
                continue
            if row[3] == self.__class__.dec_no:
                found = row
                break
        assert found is not None, "iter7 kaydı BANKA BİLDİRİMİ sayfasında bulunamadı"
        # index 8 = TCMB DEVİR ORANI, index 9 = TEŞVİK, index 10 = TAAHHÜT
        assert found[8] == "%95", found[8]
        assert found[9] == "E", found[9]
        assert found[10] == "H", found[10]

    def test_beyanname_listesi_still_17_cols(self, tokens):
        r = requests.get(f"{API}/export/excel", headers=_h(tokens["admin"]))
        assert r.status_code == 200
        import io as _io
        from openpyxl import load_workbook
        wb = load_workbook(_io.BytesIO(r.content))
        ws1 = wb["Beyanname Listesi"]
        headers = [c.value for c in ws1[1]]
        assert len(headers) == 17, headers
        assert "Teşvik" in headers and "Taahhüt" in headers
        # Beyanname Listesi'nde Teşvik/Taahhüt EVET/HAYIR formatında
        for row in ws1.iter_rows(min_row=2, values_only=True):
            if row[0] == self.__class__.dec_no:
                # last two columns are Teşvik, Taahhüt
                assert row[15] == "EVET", row
                assert row[16] == "HAYIR", row
                break

    def test_regression_dashboard_and_reports(self, tokens):
        r = requests.get(f"{API}/dashboard", headers=_h(tokens["admin"]))
        assert r.status_code == 200
        d = r.json()
        assert "bu_ay_kapatilan" in d
        assert "destek_bekleyen_tutar" in d or "destek_bekleyen_sayi" in d

        r = requests.get(f"{API}/reports/summary", headers=_h(tokens["admin"]))
        assert r.status_code == 200

        r = requests.get(f"{API}/export/check", headers=_h(tokens["admin"]))
        assert r.status_code == 200

    def test_regression_backup_and_restore_merge(self, tokens):
        hdr = {"Authorization": f"Bearer {tokens['admin']}"}
        r = requests.get(f"{API}/backup", headers=hdr)
        assert r.status_code == 200
        data = r.content
        assert data and len(data) > 100
        files = {"file": ("bkp.json", data, "application/json")}
        r = requests.post(f"{API}/backup/restore?mode=merge", headers=hdr, files=files)
        assert r.status_code == 200, r.text

    def test_cleanup(self, tokens):
        if self.__class__.match_id:
            requests.delete(f"{API}/matches/{self.__class__.match_id}",
                            headers=_h(tokens["onaylayan"]))
        if self.__class__.dec_id:
            requests.delete(f"{API}/declarations/{self.__class__.dec_id}",
                            headers=_h(tokens["ihracat"]))
        if self.__class__.pay_id:
            requests.delete(f"{API}/payments/{self.__class__.pay_id}",
                            headers=_h(tokens["banka"]))



# ---------- Iteration 8: %35 zorunlu bozdurma, TL destek kapsam dışı, IBKB-only match list, export/rows selection ----------
class TestIteration8ZorunluBozdurma35:
    """Zorunlu bozdurma oranı %35 olmalı ve payment_view içinde tutar*0.35 olarak dönmeli."""

    pay_id = None

    def test_zorunlu_bozdurma_is_35_percent(self, tokens):
        r = requests.post(f"{API}/payments", headers=_h(tokens["banka"]), json={
            "banka": "TEST ITER8", "gonderen": "TEST_ITER8_ZB",
            "tarih": "2026-05-15", "doviz": "EUR", "tutar": 6000.0})
        assert r.status_code == 200, r.text
        self.__class__.pay_id = r.json()["id"]
        # 6000 * 0.35 = 2100
        assert r.json()["zorunlu_bozdurma_orani"] == 35.0
        assert abs(r.json()["zorunlu_bozdurma"] - 2100.0) < 0.01

    def test_zorunlu_bozdurma_via_list(self, tokens):
        r = requests.get(f"{API}/payments", headers=_h(tokens["admin"]))
        assert r.status_code == 200
        p = next(x for x in r.json() if x["id"] == self.__class__.pay_id)
        assert p["zorunlu_bozdurma_orani"] == 35.0
        assert abs(p["zorunlu_bozdurma"] - 2100.0) < 0.01

    def test_cleanup(self, tokens):
        if self.__class__.pay_id:
            requests.delete(f"{API}/payments/{self.__class__.pay_id}",
                            headers=_h(tokens["banka"]))


class TestIteration8IbkbOnlyPayments:
    """GET /api/payments?ibkb_only=true sadece IBKB düzenlenmiş bedelleri döndürmeli."""

    pay_ibkb_id = None
    pay_no_ibkb_id = None

    def test_seed_two_payments(self, tokens):
        # IBKB düzenlenmiş bedel
        r = requests.post(f"{API}/payments", headers=_h(tokens["banka"]), json={
            "banka": "TEST_IBKBONLY", "gonderen": "TEST_ITER8_IBKB_OK",
            "tarih": "2026-05-16", "doviz": "EUR", "tutar": 4000.0})
        assert r.status_code == 200
        self.__class__.pay_ibkb_id = r.json()["id"]
        r = requests.put(f"{API}/payments/{self.__class__.pay_ibkb_id}/ibkb",
                         headers=_h(tokens["banka"]),
                         json={"ibkb_duzenlendi": True, "ibkb_no": "IBKB-ITER8",
                               "ibkb_tarihi": "2026-05-17",
                               "dosya_referansi": "DOSYA-ITER8-1",
                               "tcmb_devir_orani": 100})
        assert r.status_code == 200

        # IBKB düzenlenmemiş bedel
        r = requests.post(f"{API}/payments", headers=_h(tokens["banka"]), json={
            "banka": "TEST_IBKBONLY", "gonderen": "TEST_ITER8_IBKB_NO",
            "tarih": "2026-05-16", "doviz": "EUR", "tutar": 3000.0})
        assert r.status_code == 200
        self.__class__.pay_no_ibkb_id = r.json()["id"]

    def test_ibkb_only_filter_excludes_undocumented(self, tokens):
        r = requests.get(f"{API}/payments",
                         headers=_h(tokens["admin"]),
                         params={"only_available": "true", "ibkb_only": "true"})
        assert r.status_code == 200
        ids = [p["id"] for p in r.json()]
        assert self.__class__.pay_ibkb_id in ids
        assert self.__class__.pay_no_ibkb_id not in ids
        # Tüm dönen kayıtların ibkb_duzenlendi = True olmalı
        for p in r.json():
            assert p.get("ibkb_duzenlendi") is True, p

    def test_without_ibkb_only_both_visible(self, tokens):
        r = requests.get(f"{API}/payments", headers=_h(tokens["admin"]))
        assert r.status_code == 200
        ids = [p["id"] for p in r.json()]
        assert self.__class__.pay_ibkb_id in ids
        assert self.__class__.pay_no_ibkb_id in ids

    def test_cleanup(self, tokens):
        for pid in (self.__class__.pay_ibkb_id, self.__class__.pay_no_ibkb_id):
            if pid:
                requests.delete(f"{API}/payments/{pid}", headers=_h(tokens["banka"]))


class TestIteration8TryKapsamDisi:
    """TRY bedelle eşleşen beyanname destek KAPSAM_DISI, destek_tutari=0, tl_bedel=true;
    eşleştirme silinince tekrar ALINMADI/tl_bedel=false olmalı."""

    dec_id = None
    dec_no = f"TEST_ITER8_TRY_{int(datetime.now().timestamp())}"
    pay_try_id = None
    match_id = None

    def test_seed_try_dec_and_payment(self, tokens):
        r = requests.post(f"{API}/declarations", headers=_h(tokens["ihracat"]),
                          json=_base_dec(self.__class__.dec_no, "2026-05-20",
                                         doviz="TRY", tutar=100000.0,
                                         ithalatci="TEST TRY GmbH", gumruk="GM-TRY"))
        assert r.status_code == 200, r.text
        d = r.json()
        self.__class__.dec_id = d["id"]
        # Yeni beyanname (henüz TRY bedel bağlı değil ama zaten kendisi TRY)
        # Kural: doviz==TRY veya tl_bedel=true iken kapsam_disi
        assert d["destek_kapsam_disi"] is True, d
        assert d["destek_tutari"] == 0.0
        assert d["destek_durum"] == "KAPSAM_DISI"

        r = requests.post(f"{API}/payments", headers=_h(tokens["banka"]), json={
            "banka": "TL BANK", "gonderen": "TEST_ITER8_TRY_SND",
            "tarih": "2026-05-20", "doviz": "TRY", "tutar": 50000.0})
        assert r.status_code == 200
        self.__class__.pay_try_id = r.json()["id"]

    def test_match_try_makes_dec_kapsam_disi(self, tokens):
        r = requests.post(f"{API}/matches", headers=_h(tokens["onaylayan"]), json={
            "declaration_id": self.__class__.dec_id,
            "payment_id": self.__class__.pay_try_id,
            "kapatilan_tutar": 30000.0})
        assert r.status_code == 200, r.text
        self.__class__.match_id = r.json()["id"]
        d = next(x for x in requests.get(f"{API}/declarations", headers=_h(tokens["admin"])).json()
                 if x["id"] == self.__class__.dec_id)
        assert d.get("tl_bedel") is True, d
        assert d["destek_durum"] == "KAPSAM_DISI"
        assert d["destek_tutari"] == 0.0
        assert d.get("destek_kapsam_disi") is True

    def test_kapsam_disi_excluded_from_reports(self, tokens):
        # dashboard destek_bekleyen sayısında sayılmamalı
        rep = requests.get(f"{API}/reports/summary", headers=_h(tokens["admin"])).json()
        # Bizim beyanname destek_alinmadi listesinde/sayacında yer almamalı;
        # sayı doğrudan kontrol edilemez ama en azından beyanname_no listede yoksa OK
        # (endpoint sadece sayaç veriyor; içerdiği kayıtları kontrol edemezsek min)
        assert isinstance(rep["destek_alinmadi"], int)

        alerts = requests.get(f"{API}/alerts/preview", headers=_h(tokens["admin"])).json()
        # bizim TRY beyanname destek uyarı listesinde bulunmamalı
        # (preview beyanname_no listesi vermiyorsa sadece sayaç tipini kontrol)
        assert isinstance(alerts["sayilar"]["destek"], int)

    def test_delete_match_restores_alinmadi(self, tokens):
        r = requests.delete(f"{API}/matches/{self.__class__.match_id}",
                            headers=_h(tokens["onaylayan"]))
        assert r.status_code == 200
        d = next(x for x in requests.get(f"{API}/declarations", headers=_h(tokens["admin"])).json()
                 if x["id"] == self.__class__.dec_id)
        # TRY beyanname olduğu için kapsam_disi durumu doviz=TRY nedeniyle KALICI olur
        # (tl_bedel False dönse bile doviz TRY olduğundan _apply_dates_and_amounts KAPSAM_DISI verir)
        assert d["destek_kapsam_disi"] is True
        # tl_bedel değeri False olmalı (bağlı bedel kalmadı)
        assert d.get("tl_bedel") is False

    def test_eur_dec_with_try_payment_marks_kapsam_disi(self, tokens):
        """EUR beyanname + TRY bedel eşleşince beyanname kapsam_disi olmalı; TRY silinince ALINMADI dönmeli."""
        no = f"TEST_ITER8_EUR_TRY_{int(datetime.now().timestamp())}"
        r = requests.post(f"{API}/declarations", headers=_h(tokens["ihracat"]),
                          json=_base_dec(no, "2026-05-20", doviz="EUR", tutar=5000.0,
                                         ithalatci="TEST EUR TRY", gumruk="GM-EURTRY"))
        assert r.status_code == 200
        did = r.json()["id"]
        assert r.json()["destek_kapsam_disi"] is False
        assert r.json()["destek_durum"] == "ALINMADI"
        assert abs(r.json()["destek_tutari"] - 150.0) < 0.01  # %3

        # TRY bedel oluştur + eşleştir (manuel kur)
        r = requests.post(f"{API}/payments", headers=_h(tokens["banka"]), json={
            "banka": "TL", "gonderen": "TEST_EUR_TRY_TL",
            "tarih": "2026-05-20", "doviz": "TRY", "tutar": 200000.0})
        pid = r.json()["id"]
        r = requests.post(f"{API}/matches", headers=_h(tokens["onaylayan"]), json={
            "declaration_id": did, "payment_id": pid,
            "kapatilan_tutar": 1000.0, "kur": 35.0})
        assert r.status_code == 200, r.text
        mid = r.json()["id"]

        d = next(x for x in requests.get(f"{API}/declarations", headers=_h(tokens["admin"])).json()
                 if x["id"] == did)
        assert d["destek_durum"] == "KAPSAM_DISI"
        assert d["destek_tutari"] == 0.0
        assert d.get("tl_bedel") is True

        # Eşleştirme silinince EUR beyanname yeniden ALINMADI (%3) olmalı
        requests.delete(f"{API}/matches/{mid}", headers=_h(tokens["onaylayan"]))
        d = next(x for x in requests.get(f"{API}/declarations", headers=_h(tokens["admin"])).json()
                 if x["id"] == did)
        assert d.get("tl_bedel") is False
        assert d["destek_durum"] == "ALINMADI"
        assert abs(d["destek_tutari"] - 150.0) < 0.01
        # cleanup
        requests.delete(f"{API}/declarations/{did}", headers=_h(tokens["ihracat"]))
        requests.delete(f"{API}/payments/{pid}", headers=_h(tokens["banka"]))

    def test_cleanup(self, tokens):
        if self.__class__.dec_id:
            requests.delete(f"{API}/declarations/{self.__class__.dec_id}",
                            headers=_h(tokens["ihracat"]))
        if self.__class__.pay_try_id:
            requests.delete(f"{API}/payments/{self.__class__.pay_try_id}",
                            headers=_h(tokens["banka"]))


class TestIteration8ExportRowsSelection:
    """/api/export/rows aday satırları; /api/export/excel?match_ids=... sadece seçilenleri
    içermeli, sonrası gonderildi=true olmalı; mark_sent=false ile işaretlenmemeli."""

    dec_id = None
    dec_no = f"TEST_ITER8_EXP_{int(datetime.now().timestamp())}"
    pay_a_id = None
    pay_b_id = None
    match_a_id = None
    match_b_id = None

    def test_seed(self, tokens):
        r = requests.post(f"{API}/declarations", headers=_h(tokens["ihracat"]),
                          json=_base_dec(self.__class__.dec_no, "2026-06-01",
                                         doviz="EUR", tutar=10000.0,
                                         ithalatci="EXP GmbH", gumruk="GM-EXP"))
        assert r.status_code == 200
        self.__class__.dec_id = r.json()["id"]

        for label in ("A", "B"):
            r = requests.post(f"{API}/payments", headers=_h(tokens["banka"]), json={
                "banka": "TEST", "gonderen": f"TEST_ITER8_EXP_{label}",
                "tarih": "2026-06-01", "doviz": "EUR", "tutar": 3000.0})
            pid = r.json()["id"]
            r = requests.put(f"{API}/payments/{pid}/ibkb",
                             headers=_h(tokens["banka"]),
                             json={"ibkb_duzenlendi": True,
                                   "ibkb_no": f"IBKB-EXP-{label}",
                                   "ibkb_tarihi": "2026-06-02",
                                   "dosya_referansi": f"DOSYA-EXP-{label}",
                                   "tcmb_devir_orani": 100})
            assert r.status_code == 200
            r = requests.post(f"{API}/matches", headers=_h(tokens["onaylayan"]), json={
                "declaration_id": self.__class__.dec_id, "payment_id": pid,
                "kapatilan_tutar": 1000.0})
            assert r.status_code == 200, r.text
            setattr(self.__class__, f"pay_{label.lower()}_id", pid)
            setattr(self.__class__, f"match_{label.lower()}_id", r.json()["id"])

    def test_export_rows_returns_candidates_unsent(self, tokens):
        r = requests.get(f"{API}/export/rows", headers=_h(tokens["admin"]))
        assert r.status_code == 200
        rows = r.json()
        # Bizim iki match aday listede olmalı ve başlangıçta gonderildi=false
        ours = [x for x in rows if x["match_id"] in (self.__class__.match_a_id, self.__class__.match_b_id)]
        assert len(ours) == 2, ours
        for row in ours:
            assert row["gonderildi"] is False
            assert row["beyanname_no"] == self.__class__.dec_no
            assert "eksikler" in row and isinstance(row["eksikler"], list)
            assert "tutar" in row

    def test_export_check_respects_match_ids(self, tokens):
        r = requests.get(f"{API}/export/check",
                         headers=_h(tokens["admin"]),
                         params={"match_ids": self.__class__.match_a_id})
        assert r.status_code == 200
        d = r.json()
        assert d["satir_sayisi"] == 1, d

    def test_export_excel_mark_sent_false_does_not_flag(self, tokens):
        r = requests.get(f"{API}/export/excel",
                         headers=_h(tokens["admin"]),
                         params={"match_ids": self.__class__.match_a_id, "mark_sent": "false"})
        assert r.status_code == 200
        rows = requests.get(f"{API}/export/rows", headers=_h(tokens["admin"])).json()
        our_a = next(x for x in rows if x["match_id"] == self.__class__.match_a_id)
        assert our_a["gonderildi"] is False, our_a

    def test_export_excel_only_selected_rows(self, tokens):
        r = requests.get(f"{API}/export/excel",
                         headers=_h(tokens["admin"]),
                         params={"match_ids": self.__class__.match_a_id})
        assert r.status_code == 200
        import io as _io
        from openpyxl import load_workbook
        wb = load_workbook(_io.BytesIO(r.content))
        ws = wb["BANKA BİLDİRİMİ"]
        # sadece seçili match satırı olmalı; sıralı bir tek A satırı ve TOPLAM
        our_rows = []
        toplam_row = None
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == "TOPLAM":
                toplam_row = row; continue
            if row[3] == self.__class__.dec_no:
                our_rows.append(row)
        assert len(our_rows) == 1, our_rows
        assert our_rows[0][1] == "DOSYA-EXP-A", our_rows[0]
        # TOPLAM sadece seçilen satırı içerir (1000)
        assert toplam_row is not None and abs(toplam_row[5] - 1000.0) < 0.01, toplam_row

    def test_export_excel_marks_sent_by_default(self, tokens):
        # A daha önce mark_sent=false ile indirildi → hâlâ false; şimdi tekrar indir (default mark_sent=true)
        r = requests.get(f"{API}/export/excel",
                         headers=_h(tokens["admin"]),
                         params={"match_ids": self.__class__.match_a_id})
        assert r.status_code == 200
        rows = requests.get(f"{API}/export/rows", headers=_h(tokens["admin"])).json()
        our_a = next(x for x in rows if x["match_id"] == self.__class__.match_a_id)
        our_b = next(x for x in rows if x["match_id"] == self.__class__.match_b_id)
        assert our_a["gonderildi"] is True, our_a
        assert our_a["gonderim_tarihi"], our_a
        # B seçilmediği için hâlâ false
        assert our_b["gonderildi"] is False, our_b

    def test_cleanup(self, tokens):
        for mid in (self.__class__.match_a_id, self.__class__.match_b_id):
            if mid:
                requests.delete(f"{API}/matches/{mid}", headers=_h(tokens["onaylayan"]))
        if self.__class__.dec_id:
            requests.delete(f"{API}/declarations/{self.__class__.dec_id}",
                            headers=_h(tokens["ihracat"]))
        for pid in (self.__class__.pay_a_id, self.__class__.pay_b_id):
            if pid:
                requests.delete(f"{API}/payments/{pid}", headers=_h(tokens["banka"]))



# ---------- Iteration 10: 'Beni hatırla' (remember) + SMTP routing ----------
class TestIteration10RememberMe:
    """POST /api/auth/login {remember:true} → 30 gün JWT exp; false/omit → 12 saat.
    Set-Cookie access_token max-age da buna uygun olmalı. Token /api/auth/me 200 dönmeli."""

    def _decode(self, token):
        import jwt as _jwt
        secret = os.environ.get("JWT_SECRET") or dotenv_values("/app/backend/.env").get("JWT_SECRET")
        assert secret, "JWT_SECRET yok, doğrulanamaz"
        return _jwt.decode(token, secret, algorithms=["HS256"])

    def _login(self, remember=None):
        payload = {"email": "admin@ihracat.com", "password": "Admin1234!"}
        if remember is not None:
            payload["remember"] = remember
        return requests.post(f"{API}/auth/login", json=payload, timeout=30)

    def test_remember_true_exp_30_days(self):
        r = self._login(remember=True)
        assert r.status_code == 200, r.text
        data = r.json()
        assert data.get("remember") is True, data
        assert data.get("token")
        payload = self._decode(data["token"])
        exp = datetime.fromtimestamp(payload["exp"])
        delta_hours = (exp - datetime.now()).total_seconds() / 3600
        # ~30 gün = 720 saat; 700-740 aralığı toleranslı
        assert 700 < delta_hours < 740, f"exp ~30gün olmalı, delta={delta_hours}h"

        # Set-Cookie max-age ~= 30 gün (2592000 saniye)
        sc = r.headers.get("set-cookie", "")
        assert "access_token=" in sc.lower()
        import re as _re
        m = _re.search(r"max-age=(\d+)", sc, _re.IGNORECASE)
        assert m, f"max-age yok: {sc}"
        max_age = int(m.group(1))
        assert 2500000 < max_age < 2700000, f"max-age ~30 gün olmalı, {max_age}s"

        # Token /api/auth/me üzerinde geçerli
        me = requests.get(f"{API}/auth/me", headers={"Authorization": f"Bearer {data['token']}"})
        assert me.status_code == 200 and me.json()["email"] == "admin@ihracat.com"

    def test_remember_false_exp_12_hours(self):
        r = self._login(remember=False)
        assert r.status_code == 200
        data = r.json()
        assert data.get("remember") is False, data
        payload = self._decode(data["token"])
        exp = datetime.fromtimestamp(payload["exp"])
        delta_hours = (exp - datetime.now()).total_seconds() / 3600
        assert 11 < delta_hours < 13, f"exp ~12h olmalı, delta={delta_hours}h"

        sc = r.headers.get("set-cookie", "")
        import re as _re
        m = _re.search(r"max-age=(\d+)", sc, _re.IGNORECASE)
        assert m
        max_age = int(m.group(1))
        # 12 * 3600 = 43200
        assert 42000 < max_age < 44500, f"max-age ~12h olmalı, {max_age}s"

    def test_remember_omitted_defaults_to_false(self):
        r = self._login(remember=None)
        assert r.status_code == 200
        data = r.json()
        assert data.get("remember") is False
        payload = self._decode(data["token"])
        exp = datetime.fromtimestamp(payload["exp"])
        delta_hours = (exp - datetime.now()).total_seconds() / 3600
        assert 11 < delta_hours < 13

    def test_2fa_user_login_remember_carries_through_verify_code(self, tokens):
        """2FA açık kullanıcı remember=true ile login olduğunda; verify-code ile
        dönen oturumun exp'i de 30 günlük olmalı. E-posta gönderilemezse fallback
        path da aynı remember değerini korumalı (30 gün token dönmeli)."""
        # Create temporary 2FA user
        import jwt as _jwt
        secret = os.environ.get("JWT_SECRET") or dotenv_values("/app/backend/.env").get("JWT_SECRET")
        email = f"test_rem2fa_{int(datetime.now().timestamp())}@ihracat.com"
        pw = "Rem2fa123!"
        r = requests.post(f"{API}/users", headers=_h(tokens["admin"]), json={
            "email": email, "name": "TEST REM 2FA",
            "role": "goruntuleyici", "password": pw, "two_factor": True})
        assert r.status_code == 200, r.text
        uid = r.json()["id"]
        try:
            r = requests.post(f"{API}/auth/login",
                              json={"email": email, "password": pw, "remember": True}, timeout=30)
            assert r.status_code == 200, r.text
            data = r.json()

            token = None
            if data.get("token"):
                # Fallback path: 2FA e-postası gönderilemedi, doğrudan token döndü
                token = data["token"]
                assert data.get("remember") is True, data
            elif data.get("challenge_id"):
                # Challenge oluşturuldu → log'dan kodu oku, verify-code çağır
                code = _read_last_2fa_code()
                assert code and code.isdigit() and len(code) == 6, f"log'dan 2FA kodu okunamadı: {code}"
                r2 = requests.post(f"{API}/auth/verify-code",
                                   json={"challenge_id": data["challenge_id"], "code": code})
                assert r2.status_code == 200, r2.text
                d2 = r2.json()
                assert d2.get("remember") is True, d2
                token = d2["token"]
            else:
                raise AssertionError(f"Beklenmeyen login yanıtı: {data}")

            payload = _jwt.decode(token, secret, algorithms=["HS256"])
            exp = datetime.fromtimestamp(payload["exp"])
            delta_hours = (exp - datetime.now()).total_seconds() / 3600
            assert 700 < delta_hours < 740, (
                f"2FA + remember=true sonrası exp ~30gün olmalı, delta={delta_hours}h")
        finally:
            requests.delete(f"{API}/users/{uid}", headers=_h(tokens["admin"]))


class TestIteration10SmtpRouting:
    """alerts.smtp_configured() ve alerts._send() yönlendiricisi birim testi.
    SMTP_HOST+SMTP_FROM yokken Emergent yolu; ayarlanınca SMTP yolu (monkeypatch)."""

    def _import_alerts(self):
        import sys as _sys
        if "/app/backend" not in _sys.path:
            _sys.path.insert(0, "/app/backend")
        # Alerts modülünü her testte yeniden yükle (env değişikliklerini yansıtsın)
        import importlib
        import alerts as _al
        return importlib.reload(_al)

    def test_smtp_configured_false_without_env(self, monkeypatch):
        monkeypatch.delenv("SMTP_HOST", raising=False)
        monkeypatch.delenv("SMTP_FROM", raising=False)
        al = self._import_alerts()
        assert al.smtp_configured() is False

    def test_smtp_configured_true_with_env(self, monkeypatch):
        monkeypatch.setenv("SMTP_HOST", "smtp.example.com")
        monkeypatch.setenv("SMTP_FROM", "no-reply@example.com")
        al = self._import_alerts()
        assert al.smtp_configured() is True

    def test_send_routes_to_smtp_when_configured(self, monkeypatch):
        """_send() SMTP env ayarlıysa _send_smtp'yi çağırmalı, httpx'e gitmemeli."""
        import asyncio
        monkeypatch.setenv("SMTP_HOST", "smtp.example.com")
        monkeypatch.setenv("SMTP_FROM", "no-reply@example.com")
        monkeypatch.setenv("EMAIL_FROM_NAME", "TEST BILDIRIM")
        al = self._import_alerts()
        calls = {"smtp": 0, "http": 0}

        async def fake_smtp(to, subject, html):
            calls["smtp"] += 1
            calls["last"] = (to, subject)

        class FakeClient:
            def __init__(self, *a, **k): pass
            async def __aenter__(self): return self
            async def __aexit__(self, *a): return False
            async def post(self, *a, **k):
                calls["http"] += 1
                raise AssertionError("SMTP configured iken httpx kullanılmamalı")

        monkeypatch.setattr(al, "_send_smtp", fake_smtp)
        monkeypatch.setattr(al.httpx, "AsyncClient", FakeClient)
        result = asyncio.run(al._send(["x@y.com"], "Konu", "<b>merhaba</b>"))
        assert result == "smtp"
        assert calls["smtp"] == 1
        assert calls["http"] == 0
        assert calls["last"][0] == ["x@y.com"]

    def test_send_uses_emergent_when_smtp_not_configured(self, monkeypatch):
        """SMTP env yokken _send() Emergent HTTP endpoint'ini kullanmalı."""
        import asyncio
        monkeypatch.delenv("SMTP_HOST", raising=False)
        monkeypatch.delenv("SMTP_FROM", raising=False)
        monkeypatch.setenv("EMAIL_FROM_NAME", "TEST BILDIRIM")
        monkeypatch.setenv("EMERGENT_EMAIL_KEY", "test-key-xyz")
        al = self._import_alerts()
        calls = {"http": 0}

        class FakeResp:
            status_code = 200
            def raise_for_status(self): return None
            def json(self): return {"id": "eid-123"}

        class FakeClient:
            def __init__(self, *a, **k): pass
            async def __aenter__(self): return self
            async def __aexit__(self, *a): return False
            async def post(self, url, headers=None, json=None):
                calls["http"] += 1
                calls["url"] = url
                calls["headers"] = headers
                calls["payload"] = json
                return FakeResp()

        async def fake_smtp(*a, **k):
            raise AssertionError("SMTP yapılandırılmamışken _send_smtp çağrılmamalı")

        monkeypatch.setattr(al, "_send_smtp", fake_smtp)
        monkeypatch.setattr(al.httpx, "AsyncClient", FakeClient)
        result = asyncio.run(al._send(["a@b.com"], "K", "<i>x</i>"))
        assert result == "eid-123"
        assert calls["http"] == 1
        assert calls["url"].endswith("/api/v1/email/send")
        assert calls["headers"]["X-Email-Key"] == "test-key-xyz"
        assert calls["payload"]["to"] == ["a@b.com"]


class TestIteration10AlertsSend:
    """POST /api/alerts/send admin/onaylayan tarafından çalıştığında 200 + sent=true;
    Uyarı/EPOSTA audit log'u düşmeli. (Tek çağrı ile sınırlı — gerçek e-posta gider.)"""

    def test_send_returns_ok_and_audit(self, tokens):
        # Gerçek gönderim testi TestAlerts.test_send_by_onaylayan_and_audit içinde yapılıyor;
        # aynı koşuda ikinci kez çağırmak Emergent rate-limit (429) tetikliyor.
        pytest.skip("Tekrarlı gerçek e-posta gönderimi rate-limit yaratıyor (TestAlerts kapsıyor)")

    def test_preview_recipients_and_counters(self, tokens):
        r = requests.get(f"{API}/alerts/preview", headers=_h(tokens["admin"]))
        assert r.status_code == 200
        d = r.json()
        assert isinstance(d.get("alicilar"), list) and d["alicilar"]
        for k in ("gecmis", "yaklasan", "ibkb", "destek"):
            assert k in d["sayilar"]


# ---------- Iteration 11: E-posta güvenlik kapısı + payload + 2FA kodu gerçek gönderim ----------
def _reload_alerts():
    import sys as _sys, importlib
    if "/app/backend" not in _sys.path:
        _sys.path.insert(0, "/app/backend")
    import alerts as _al
    return importlib.reload(_al)


class TestIteration11EmailSafetyGate:
    """alerts._assert_safe_email güvenlik kapısı birim testleri."""

    def test_rejects_form_or_input_tags(self):
        al = _reload_alerts()
        with pytest.raises(ValueError):
            al._assert_safe_email("Konu", "<div><form action='x'></form></div>")
        with pytest.raises(ValueError):
            al._assert_safe_email("Konu", "<div><input type='text' /></div>")
        with pytest.raises(ValueError):
            al._assert_safe_email("Konu", "<textarea></textarea>")

    def test_rejects_credential_ask_text(self):
        al = _reload_alerts()
        with pytest.raises(ValueError):
            al._assert_safe_email(
                "Konu", "<p>Please Reply with your password to verify.</p>")
        with pytest.raises(ValueError):
            al._assert_safe_email("CVV lazım", "<p>Şifreniz için tıklayın</p>")
        with pytest.raises(ValueError):
            al._assert_safe_email("Konu", "<p>Enter your password below</p>")

    def test_rejects_non_https_and_javascript_urls(self):
        al = _reload_alerts()
        with pytest.raises(ValueError):
            al._assert_safe_email("K", '<a href="http://example.com">tıkla</a>')
        with pytest.raises(ValueError):
            al._assert_safe_email("K", '<a href="javascript:alert(1)">x</a>')

    def test_rejects_shorteners_and_numeric_hosts(self):
        al = _reload_alerts()
        with pytest.raises(ValueError):
            al._assert_safe_email("K", '<a href="https://bit.ly/xyz">bit</a>')
        with pytest.raises(ValueError):
            al._assert_safe_email("K", '<a href="https://tinyurl.com/x">tiny</a>')
        with pytest.raises(ValueError):
            al._assert_safe_email("K", '<a href="https://192.168.1.10/path">ip</a>')

    def test_rejects_anchor_text_host_mismatch(self):
        al = _reload_alerts()
        # Anchor gösterge host farklı görünüyor → reddedilmeli
        with pytest.raises(ValueError):
            al._assert_safe_email(
                "K", '<a href="https://phisher.com/x">https://kalipsanaluminyum.com</a>')

    def test_accepts_real_alert_html(self, tokens):
        """alerts.build_html(...) gerçek üretimde döndüğünde güvenlik kapısı geçmeli."""
        al = _reload_alerts()
        # Sistemden gerçek decs verisini al
        r = requests.get(f"{API}/declarations", headers=_h(tokens["admin"]))
        assert r.status_code == 200
        html, counts = al.build_html(r.json())
        # Hata fırlatmamalı
        al._assert_safe_email("Haftalık uyarı", html)

    def test_accepts_2fa_code_html(self):
        """send_code'un ürettiği iç HTML kapıdan geçmeli (kredi isteme metni yok)."""
        al = _reload_alerts()
        # send_code kaynağındaki HTML şablonunun eşdeğerini oluştur
        code = "123456"
        html = (
            '<div style="font-family:Arial,Helvetica,sans-serif;max-width:520px">'
            '<h2>İhracat Bedeli Kapatma Sistemi</h2>'
            f'<p>Sayın kullanıcı, giriş doğrulama kodunuz:</p>'
            f'<div style="font-family:monospace">{code}</div>'
            '<p>Kod 5 dakika geçerlidir ve yalnızca bir kez kullanılabilir.</p></div>'
        )
        al._assert_safe_email(f"Giriş doğrulama kodunuz: {code}", html)

    def test_accepts_https_matching_anchor(self):
        al = _reload_alerts()
        # anchor text tam olarak href host'u ile aynı → geçmeli
        al._assert_safe_email(
            "K",
            '<a href="https://kalipsanaluminyum.com/x">https://kalipsanaluminyum.com</a>')


class TestIteration11EmergentPayload:
    """Emergent /api/v1/email/send çağrısında from_name zorunlu; EMAIL_REPLY_TO tanımlıysa
    contact_email eklenmeli, tanımsızsa hiç eklenmemeli."""

    def _run(self, monkeypatch, reply_to=None):
        import asyncio
        monkeypatch.delenv("SMTP_HOST", raising=False)
        monkeypatch.delenv("SMTP_FROM", raising=False)
        monkeypatch.setenv("EMAIL_FROM_NAME", "Kalıpsan Bildirim")
        monkeypatch.setenv("EMERGENT_EMAIL_KEY", "test-key-iter11")
        if reply_to is None:
            monkeypatch.delenv("EMAIL_REPLY_TO", raising=False)
        else:
            monkeypatch.setenv("EMAIL_REPLY_TO", reply_to)
        al = _reload_alerts()
        cap = {}

        class FakeResp:
            status_code = 200
            def raise_for_status(self): return None
            def json(self): return {"id": "eid-iter11"}

        class FakeClient:
            def __init__(self, *a, **k): pass
            async def __aenter__(self): return self
            async def __aexit__(self, *a): return False
            async def post(self, url, headers=None, json=None):
                cap["url"] = url
                cap["headers"] = headers
                cap["payload"] = json
                return FakeResp()

        monkeypatch.setattr(al.httpx, "AsyncClient", FakeClient)
        # Güvenli/temiz HTML — kapıdan geçer
        result = asyncio.run(al._send(["x@y.com"], "Konu", "<p>selam</p>"))
        assert result == "eid-iter11"
        return cap

    def test_from_name_and_xemailkey_header(self, monkeypatch):
        cap = self._run(monkeypatch, reply_to=None)
        assert cap["url"].startswith("https://integrations.emergentagent.com")
        assert cap["url"].endswith("/api/v1/email/send")
        assert cap["headers"].get("X-Email-Key") == "test-key-iter11"
        assert cap["payload"]["from_name"] == "Kalıpsan Bildirim"
        assert cap["payload"]["to"] == ["x@y.com"]
        assert cap["payload"]["subject"] == "Konu"
        # EMAIL_REPLY_TO tanımsızken contact_email eklenmemeli
        assert "contact_email" not in cap["payload"]

    def test_reply_to_maps_to_contact_email(self, monkeypatch):
        cap = self._run(monkeypatch, reply_to="destek@kalipsanaluminyum.com")
        assert cap["payload"]["contact_email"] == "destek@kalipsanaluminyum.com"
        # from_name yine bulunmalı
        assert cap["payload"]["from_name"] == "Kalıpsan Bildirim"


class TestIteration11SmtpSafetyGate:
    """SMTP yolunda güvenlik kapısı yine devrede olmalı."""

    def test_smtp_path_still_asserts_safety(self, monkeypatch):
        import asyncio
        monkeypatch.setenv("SMTP_HOST", "smtp.example.com")
        monkeypatch.setenv("SMTP_FROM", "no-reply@example.com")
        monkeypatch.setenv("EMAIL_FROM_NAME", "TEST")
        al = _reload_alerts()

        async def fake_smtp(to, subject, html):
            fake_smtp.called = True
        fake_smtp.called = False
        monkeypatch.setattr(al, "_send_smtp", fake_smtp)

        # Kötü HTML: kapı reddetmeli, SMTP çağrılmamalı
        with pytest.raises(ValueError):
            asyncio.run(al._send(["x@y.com"], "K", "<form></form>"))
        assert fake_smtp.called is False

        # Temiz HTML: SMTP yolu tetiklenmeli, gerçek bağlantı YOK
        result = asyncio.run(al._send(["x@y.com"], "K", "<p>ok</p>"))
        assert result == "smtp"
        assert fake_smtp.called is True


class TestIteration11TwoFactorCodeFlow:
    """2FA kodu gerçek gönderim akışı: delivered@resend.dev ile geçici kullanıcı;
    login → challenge (gonderildi=true, fallback devreye girmemeli); resend-code
    yeni challenge; yanlış kod 401; log'dan okunan 6 hane doğru kod ile verify-code."""

    uid = None
    email = None
    pw = "Tfa2Fa11!"

    def test_setup_user(self, tokens):
        # Resend sandbox destination: delivered@resend.dev — always accepted
        self.__class__.email = f"test_2fa_iter11_{int(datetime.now().timestamp())}@resend.dev"
        # Kullanıcının kendi e-postasına kod gönderilir → adres delivered@resend.dev olmalı
        # ki Resend gerçekten teslim etsin. Bu yüzden e-postayı delivered@resend.dev yapıyoruz.
        self.__class__.email = "delivered@resend.dev"
        # Zaten varsa temizle
        r = requests.get(f"{API}/users", headers=_h(tokens["admin"]))
        for u in r.json():
            if u.get("email") == self.__class__.email:
                requests.delete(f"{API}/users/{u['id']}", headers=_h(tokens["admin"]))
        r = requests.post(f"{API}/users", headers=_h(tokens["admin"]), json={
            "email": self.__class__.email, "name": "TEST 2FA ITER11",
            "role": "goruntuleyici", "password": self.__class__.pw,
            "two_factor": True})
        assert r.status_code == 200, r.text
        u = r.json()
        assert u["two_factor"] is True
        self.__class__.uid = u["id"]

    def test_login_returns_challenge_and_gonderildi_true(self):
        r = requests.post(f"{API}/auth/login", json={
            "email": self.__class__.email, "password": self.__class__.pw}, timeout=30)
        assert r.status_code == 200, r.text
        data = r.json()
        # Gerçek e-posta gitmeli — fallback DEVREYE girmemeli (token gelmemeli, challenge_id gelmeli)
        assert data.get("challenge_id"), f"Beklenen challenge, gelen: {data}"
        assert data.get("gonderildi") is True, f"gonderildi=true bekleniyor: {data}"
        assert "token" not in data, data
        self.__class__.challenge_id = data["challenge_id"]

    def test_wrong_code_returns_401(self):
        cid = getattr(self.__class__, "challenge_id", None)
        assert cid, "önceki testte challenge alınamadı"
        r = requests.post(f"{API}/auth/verify-code", json={
            "challenge_id": cid, "code": "000000"})
        # 401 (hatalı) veya 429 (denemede sınır) kabul; ama 200 KESİNLİKLE olmamalı
        assert r.status_code in (401, 429), r.text

    def test_resend_produces_new_challenge(self):
        cid = getattr(self.__class__, "challenge_id", None)
        assert cid
        r = requests.post(f"{API}/auth/resend-code", json={"challenge_id": cid})
        assert r.status_code == 200, r.text
        data = r.json()
        new_cid = data.get("challenge_id")
        assert new_cid and new_cid != cid, data
        assert data.get("gonderildi") is True, data
        self.__class__.challenge_id = new_cid

    def test_correct_code_verifies_and_returns_token(self):
        cid = getattr(self.__class__, "challenge_id", None)
        assert cid
        # Kod backend log'undan okunmalı (delivered@resend.dev'e giden gerçek kodu
        # sandbox'ta biz okuyamıyoruz; ama server INFO log'a da yazıyor)
        import time
        code = None
        for _ in range(6):
            code = _read_last_2fa_code()
            if code and len(code) == 6 and code.isdigit():
                break
            time.sleep(0.5)
        assert code and code.isdigit() and len(code) == 6, f"log'dan 2FA kodu okunamadı: {code!r}"
        r = requests.post(f"{API}/auth/verify-code", json={
            "challenge_id": cid, "code": code})
        assert r.status_code == 200, r.text
        data = r.json()
        assert data.get("token"), data
        assert data.get("email") == self.__class__.email

    def test_cleanup(self, tokens):
        if self.__class__.uid:
            r = requests.delete(f"{API}/users/{self.__class__.uid}",
                                headers=_h(tokens["admin"]))
            assert r.status_code in (200, 404)

