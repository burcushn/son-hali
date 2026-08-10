from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

import os
import io
import logging
from datetime import datetime, timezone, timedelta

import bcrypt
import jwt
import secrets
from bson import ObjectId
from fastapi import FastAPI, APIRouter, HTTPException, Request, Response, Depends
from fastapi.responses import StreamingResponse
from motor.motor_asyncio import AsyncIOMotorClient
from starlette.middleware.cors import CORSMiddleware
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from models import (
    User, UserCreate, UserUpdate, LoginInput, VerifyCodeInput, ResendCodeInput,
    Declaration, DeclarationInput, Payment, PaymentInput, IbkbInput,
    Match, MatchInput, ROLES, ROLE_LABELS, utcnow_iso,
)
import tcmb
import alerts
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

client = AsyncIOMotorClient(os.environ["MONGO_URL"])
db = client[os.environ["DB_NAME"]]

app = FastAPI()
api = APIRouter(prefix="/api")

JWT_ALG = "HS256"


def hash_password(p: str) -> str:
    return bcrypt.hashpw(p.encode(), bcrypt.gensalt()).decode()


def verify_password(p: str, h: str) -> bool:
    try:
        return bcrypt.checkpw(p.encode(), h.encode())
    except Exception:
        return False


def create_access_token(uid: str, email: str) -> str:
    return jwt.encode(
        {"sub": uid, "email": email, "type": "access",
         "exp": datetime.now(timezone.utc) + timedelta(hours=12)},
        os.environ["JWT_SECRET"], algorithm=JWT_ALG)


async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth = request.headers.get("Authorization", "")
        if auth.startswith("Bearer "):
            token = auth[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Oturum bulunamadı")
    try:
        payload = jwt.decode(token, os.environ["JWT_SECRET"], algorithms=[JWT_ALG])
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Oturum süresi doldu")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Geçersiz oturum")
    user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
    if not user or not user.get("active", True):
        raise HTTPException(status_code=401, detail="Kullanıcı bulunamadı veya pasif")
    user["_id"] = str(user["_id"])
    user.pop("password_hash", None)
    return user


def require(*roles):
    async def dep(user: dict = Depends(get_current_user)):
        if user["role"] != "admin" and user["role"] not in roles:
            raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
        return user
    return dep


async def log_action(user: dict, modul: str, islem: str, aciklama: str, ref: str = ""):
    await db.audit_logs.insert_one({
        "modul": modul, "islem": islem, "aciklama": aciklama, "ref": ref,
        "kullanici_ad": user.get("name", ""), "kullanici_rol": user.get("role", ""),
        "tarih": utcnow_iso(),
    })


# ---------------- helpers ----------------

def declaration_view(doc: dict) -> dict:
    d = Declaration.from_mongo(doc).model_dump()
    d["kalan"] = round(d["tutar"] - d["kapatilan"], 2)
    d["destek_tutari"] = round(d["tutar"] * 0.03, 2)
    d["ibkb_durum"] = "DUZENLENDI" if d.get("ibkb_alindi") else "DUZENLENMEDI"
    d["destek_durum"] = "ALINDI" if d.get("destek_alindi") else "ALINMADI"
    base = d.get("kapanis_tarihi") or d.get("acilis_tarihi") or ""
    try:
        son = datetime.strptime(base[:10], "%Y-%m-%d") + timedelta(days=180)
        d["son_kapatma_tarihi"] = son.strftime("%Y-%m-%d")
        kalan_gun = (son.date() - datetime.now().date()).days
    except Exception:
        d["son_kapatma_tarihi"] = ""
        kalan_gun = None
    d["kalan_gun"] = kalan_gun
    if d["durum"] == "KAPALI":
        d["sure_durum"] = "TAMAM"
    elif kalan_gun is None:
        d["sure_durum"] = "NORMAL"
    elif kalan_gun < 0:
        d["sure_durum"] = "GECMIS"
    elif kalan_gun <= 30:
        d["sure_durum"] = "YAKLASAN"
    else:
        d["sure_durum"] = "NORMAL"
    return d


_LAST_ACH: dict = {}


def payment_view(doc: dict) -> dict:
    p = Payment.from_mongo(doc).model_dump()
    p["bakiye"] = round(p["tutar"] - p["kullanilan"], 2)
    p["ibkb_durum"] = "DUZENLENDI" if p.get("ibkb_duzenlendi") else "DUZENLENMEDI"
    p["zorunlu_bozdurma"] = round(p["tutar"] * 0.30, 2)
    p["ach_iban_default"] = _LAST_ACH.get("iban") or os.environ.get("DEFAULT_ACH_IBAN", "")
    return p


def dstatus(tutar: float, kapatilan: float) -> str:
    if kapatilan <= 0.004:
        return "ACIK"
    if kapatilan >= tutar - 0.01:
        return "KAPALI"
    return "KISMI"


def pstatus(tutar: float, kullanilan: float) -> str:
    if kullanilan <= 0.004:
        return "KULLANILMADI"
    if kullanilan >= tutar - 0.01:
        return "TUKENDI"
    return "KISMI"


async def recalc_declaration(did: str):
    ms = await db.matches.find({"declaration_id": did}).to_list(2000)
    total = round(sum(m["kapatilan_tutar"] for m in ms), 2)
    doc = await db.declarations.find_one({"_id": ObjectId(did)})
    if doc:
        await db.declarations.update_one({"_id": ObjectId(did)},
            {"$set": {"kapatilan": total, "durum": dstatus(doc["tutar"], total)}})


async def recalc_payment(pid: str):
    ms = await db.matches.find({"payment_id": pid}).to_list(2000)
    total = round(sum(m["bedel_kullanilan"] for m in ms), 2)
    doc = await db.payments.find_one({"_id": ObjectId(pid)})
    if doc:
        await db.payments.update_one({"_id": ObjectId(pid)},
            {"$set": {"kullanilan": total, "durum": pstatus(doc["tutar"], total)}})


# ---------------- auth ----------------
def _issue_session(user: dict, response: Response) -> dict:
    token = create_access_token(str(user["_id"]), user["email"])
    response.set_cookie("access_token", token, httponly=True, secure=True,
                        samesite="none", max_age=43200, path="/")
    u = dict(user)
    u["_id"] = str(u["_id"])
    u.pop("password_hash", None)
    return {**User.from_mongo(u).model_dump(), "token": token}


async def _create_challenge(user: dict) -> tuple:
    code = f"{secrets.randbelow(1000000):06d}"
    res = await db.login_challenges.insert_one({
        "user_id": str(user["_id"]), "email": user["email"],
        "code_hash": hash_password(code), "attempts": 0,
        "expires_at": datetime.now(timezone.utc) + timedelta(minutes=5),
        "used": False, "created_at": utcnow_iso(),
    })
    logger.info(f"[2FA] {user['email']} doğrulama kodu: {code}")
    delivered = await alerts.send_code(user["email"], user.get("name", ""), code)
    return str(res.inserted_id), delivered


@api.post("/auth/login")
async def login(body: LoginInput, response: Response):
    email = body.email.strip().lower()
    user = await db.users.find_one({"email": email})
    if not user or not verify_password(body.password, user.get("password_hash", "")):
        if user:
            fails = user.get("failed_logins", 0) + 1
            upd = {"failed_logins": fails}
            if fails >= 5:
                upd["locked_until"] = (datetime.now(timezone.utc) + timedelta(minutes=15)).isoformat()
            await db.users.update_one({"_id": user["_id"]}, {"$set": upd})
        raise HTTPException(status_code=401, detail="E-posta veya şifre hatalı")
    if not user.get("active", True):
        raise HTTPException(status_code=403, detail="Hesabınız pasif durumda")
    locked = user.get("locked_until")
    if locked and datetime.fromisoformat(locked) > datetime.now(timezone.utc):
        raise HTTPException(status_code=429,
                            detail="Çok fazla hatalı giriş denemesi. 15 dakika sonra tekrar deneyin.")
    await db.users.update_one({"_id": user["_id"]},
                              {"$set": {"failed_logins": 0, "locked_until": None}})
    if not user.get("two_factor", False):
        return _issue_session(user, response)
    cid, delivered = await _create_challenge(user)
    return {"two_factor": True, "challenge_id": cid, "email": email, "gonderildi": delivered,
            "mesaj": ("Doğrulama kodu e-posta adresinize gönderildi (5 dakika geçerli)."
                      if delivered else
                      "Kod oluşturuldu ancak e-posta gönderilemedi. 'Kodu tekrar gönder' ile "
                      "yeniden deneyin veya yöneticinizle iletişime geçin.")}


@api.post("/auth/verify-code")
async def verify_code(body: VerifyCodeInput, response: Response):
    ch = await db.login_challenges.find_one({"_id": ObjectId(body.challenge_id)})
    if not ch or ch.get("used"):
        raise HTTPException(status_code=400, detail="Doğrulama isteği geçersiz, tekrar giriş yapın")
    exp = ch["expires_at"]
    if exp.tzinfo is None:
        exp = exp.replace(tzinfo=timezone.utc)
    if exp < datetime.now(timezone.utc):
        raise HTTPException(status_code=400, detail="Kodun süresi doldu, yeni kod isteyin")
    if ch.get("attempts", 0) >= 5:
        raise HTTPException(status_code=429, detail="Çok fazla hatalı deneme, yeni kod isteyin")
    if not verify_password(body.code.strip(), ch["code_hash"]):
        await db.login_challenges.update_one({"_id": ch["_id"]}, {"$inc": {"attempts": 1}})
        kalan = 4 - ch.get("attempts", 0)
        if kalan <= 0:
            raise HTTPException(status_code=429, detail="Çok fazla hatalı deneme, yeni kod isteyin")
        raise HTTPException(status_code=401, detail=f"Doğrulama kodu hatalı ({kalan} deneme kaldı)")
    await db.login_challenges.update_one({"_id": ch["_id"]}, {"$set": {"used": True}})
    user = await db.users.find_one({"_id": ObjectId(ch["user_id"])})
    if not user or not user.get("active", True):
        raise HTTPException(status_code=403, detail="Hesabınız pasif durumda")
    return _issue_session(user, response)


@api.post("/auth/resend-code")
async def resend_code(body: ResendCodeInput):
    ch = await db.login_challenges.find_one({"_id": ObjectId(body.challenge_id)})
    if not ch or ch.get("used"):
        raise HTTPException(status_code=400, detail="Doğrulama isteği geçersiz, tekrar giriş yapın")
    user = await db.users.find_one({"_id": ObjectId(ch["user_id"])})
    if not user:
        raise HTTPException(status_code=404, detail="Kullanıcı bulunamadı")
    await db.login_challenges.update_one({"_id": ch["_id"]}, {"$set": {"used": True}})
    cid, delivered = await _create_challenge(user)
    return {"challenge_id": cid, "gonderildi": delivered,
            "mesaj": ("Yeni doğrulama kodu gönderildi." if delivered
                      else "Kod oluşturuldu ancak e-posta gönderilemedi, tekrar deneyin.")}


@api.post("/auth/logout")
async def logout(response: Response, user: dict = Depends(get_current_user)):
    response.delete_cookie("access_token", path="/")
    return {"ok": True}


@api.get("/auth/me")
async def me(user: dict = Depends(get_current_user)):
    return User.from_mongo(user).model_dump()


# ---------------- users ----------------
@api.get("/users")
async def list_users(user: dict = Depends(require())):
    docs = await db.users.find({}, {"password_hash": 0}).sort("created_at", 1).to_list(500)
    return [User.from_mongo(d).model_dump() for d in docs]


@api.post("/users")
async def create_user(body: UserCreate, user: dict = Depends(require())):
    if body.role not in ROLES:
        raise HTTPException(status_code=400, detail="Geçersiz rol")
    email = body.email.strip().lower()
    if await db.users.find_one({"email": email}):
        raise HTTPException(status_code=400, detail="Bu e-posta zaten kayıtlı")
    doc = {"email": email, "name": body.name, "role": body.role, "active": True,
           "two_factor": body.two_factor,
           "password_hash": hash_password(body.password), "created_at": utcnow_iso()}
    res = await db.users.insert_one(doc)
    await log_action(user, "Kullanıcı", "EKLE", f"{body.name} ({ROLE_LABELS[body.role]}) eklendi", str(res.inserted_id))
    doc["_id"] = res.inserted_id
    return User.from_mongo(doc).model_dump()


@api.put("/users/{uid}")
async def update_user(uid: str, body: UserUpdate, user: dict = Depends(require())):
    upd = {k: v for k, v in body.model_dump(exclude_none=True).items() if k != "password"}
    if body.password:
        upd["password_hash"] = hash_password(body.password)
    if not upd:
        raise HTTPException(status_code=400, detail="Güncellenecek alan yok")
    await db.users.update_one({"_id": ObjectId(uid)}, {"$set": upd})
    await log_action(user, "Kullanıcı", "GUNCELLE", f"Kullanıcı güncellendi", uid)
    doc = await db.users.find_one({"_id": ObjectId(uid)}, {"password_hash": 0})
    return User.from_mongo(doc).model_dump()


@api.delete("/users/{uid}")
async def delete_user(uid: str, user: dict = Depends(require())):
    if uid == user["_id"]:
        raise HTTPException(status_code=400, detail="Kendi hesabınızı silemezsiniz")
    await db.users.delete_one({"_id": ObjectId(uid)})
    await log_action(user, "Kullanıcı", "SIL", "Kullanıcı silindi", uid)
    return {"ok": True}


# ---------------- declarations ----------------
@api.get("/declarations")
async def list_declarations(durum: str = "", q: str = "", sure: str = "",
                            ibkb: str = "", destek: str = "",
                            user: dict = Depends(get_current_user)):
    query = {}
    if durum:
        query["durum"] = durum
    if q:
        query["$or"] = [{"beyanname_no": {"$regex": q, "$options": "i"}},
                        {"ithalatci": {"$regex": q, "$options": "i"}},
                        {"gumruk_mudurlugu_no": {"$regex": q, "$options": "i"}}]
    docs = await db.declarations.find(query).sort("acilis_tarihi", -1).to_list(2000)
    items = [declaration_view(d) for d in docs]
    if sure:
        items = [i for i in items if i["sure_durum"] == sure]
    if ibkb:
        items = [i for i in items if i["ibkb_durum"] == ibkb]
    if destek:
        items = [i for i in items if i["destek_durum"] == destek]
    return items


@api.post("/declarations")
async def create_declaration(body: DeclarationInput, user: dict = Depends(require("ihracat"))):
    if await db.declarations.find_one({"beyanname_no": body.beyanname_no.strip()}):
        raise HTTPException(status_code=400, detail="Bu beyanname numarası zaten kayıtlı")
    d = Declaration(**body.model_dump(), created_by=user["name"])
    d.beyanname_no = d.beyanname_no.strip()
    d.durum = "ACIK"
    res = await db.declarations.insert_one(d.to_mongo())
    await log_action(user, "Beyanname", "EKLE",
                     f"{d.beyanname_no} / {d.tutar} {d.doviz} eklendi", str(res.inserted_id))
    doc = await db.declarations.find_one({"_id": res.inserted_id})
    return declaration_view(doc)


@api.put("/declarations/{did}")
async def update_declaration(did: str, body: DeclarationInput, user: dict = Depends(require("ihracat"))):
    doc = await db.declarations.find_one({"_id": ObjectId(did)})
    if not doc:
        raise HTTPException(status_code=404, detail="Beyanname bulunamadı")
    if body.tutar < doc.get("kapatilan", 0) - 0.01:
        raise HTTPException(status_code=400, detail="Yeni tutar, kapatılan tutardan küçük olamaz")
    upd = body.model_dump()
    await db.declarations.update_one({"_id": ObjectId(did)},
        {"$set": {**upd, "durum": dstatus(body.tutar, doc.get("kapatilan", 0))}})
    await log_action(user, "Beyanname", "GUNCELLE", f"{body.beyanname_no} güncellendi", did)
    return declaration_view(await db.declarations.find_one({"_id": ObjectId(did)}))


@api.delete("/declarations/{did}")
async def delete_declaration(did: str, user: dict = Depends(require("ihracat"))):
    if await db.matches.find_one({"declaration_id": did}):
        raise HTTPException(status_code=400, detail="Eşleştirmesi olan beyanname silinemez. Önce eşleştirmeleri geri alın.")
    doc = await db.declarations.find_one({"_id": ObjectId(did)})
    await db.declarations.delete_one({"_id": ObjectId(did)})
    await log_action(user, "Beyanname", "SIL", f"{doc.get('beyanname_no','')} silindi", did)
    return {"ok": True}


# ---------------- payments ----------------
@api.get("/payments")
async def list_payments(durum: str = "", q: str = "", only_available: bool = False,
                        user: dict = Depends(get_current_user)):
    query = {}
    if durum:
        query["durum"] = durum
    if q:
        query["$or"] = [{"gonderen": {"$regex": q, "$options": "i"}},
                        {"banka": {"$regex": q, "$options": "i"}},
                        {"aciklama": {"$regex": q, "$options": "i"}}]
    docs = await db.payments.find(query).sort("tarih", -1).to_list(2000)
    items = [payment_view(d) for d in docs]
    if only_available:
        items = [i for i in items if i["bakiye"] > 0.01]
    return items


@api.post("/payments")
async def create_payment(body: PaymentInput, user: dict = Depends(require("banka"))):
    p = Payment(**body.model_dump(), created_by=user["name"])
    res = await db.payments.insert_one(p.to_mongo())
    await log_action(user, "Bedel", "EKLE",
                     f"{p.gonderen} / {p.tutar} {p.doviz} ({p.banka}) eklendi", str(res.inserted_id))
    return payment_view(await db.payments.find_one({"_id": res.inserted_id}))


@api.put("/payments/{pid}")
async def update_payment(pid: str, body: PaymentInput, user: dict = Depends(require("banka"))):
    doc = await db.payments.find_one({"_id": ObjectId(pid)})
    if not doc:
        raise HTTPException(status_code=404, detail="Bedel bulunamadı")
    if body.tutar < doc.get("kullanilan", 0) - 0.01:
        raise HTTPException(status_code=400, detail="Yeni tutar, kullanılan tutardan küçük olamaz")
    await db.payments.update_one({"_id": ObjectId(pid)},
        {"$set": {**body.model_dump(), "durum": pstatus(body.tutar, doc.get("kullanilan", 0))}})
    await log_action(user, "Bedel", "GUNCELLE", f"{body.gonderen} bedeli güncellendi", pid)
    return payment_view(await db.payments.find_one({"_id": ObjectId(pid)}))


@api.delete("/payments/{pid}")
async def delete_payment(pid: str, user: dict = Depends(require("banka"))):
    if await db.matches.find_one({"payment_id": pid}):
        raise HTTPException(status_code=400, detail="Eşleştirmesi olan bedel silinemez. Önce eşleştirmeleri geri alın.")
    doc = await db.payments.find_one({"_id": ObjectId(pid)})
    await db.payments.delete_one({"_id": ObjectId(pid)})
    await log_action(user, "Bedel", "SIL", f"{doc.get('gonderen','')} bedeli silindi", pid)
    return {"ok": True}


# ---------------- IBKB ----------------
@api.put("/payments/{pid}/ibkb")
async def update_ibkb(pid: str, body: IbkbInput, user: dict = Depends(require("banka"))):
    p = await db.payments.find_one({"_id": ObjectId(pid)})
    if not p:
        raise HTTPException(status_code=404, detail="Bedel bulunamadı")
    upd = body.model_dump(exclude_none=True)
    upd["ibkb_duzenlendi"] = True
    if upd.get("ach_iban"):
        _LAST_ACH["iban"] = upd["ach_iban"]
    await db.payments.update_one({"_id": ObjectId(pid)}, {"$set": upd})
    await log_action(user, "IBKB", "GUNCELLE",
        f"{p['gonderen']} bedeli için IBKB bilgileri kaydedildi "
        f"(IBKB No: {body.ibkb_no or '-'}, Dosya Ref: {body.dosya_referansi or '-'})", pid)
    return payment_view(await db.payments.find_one({"_id": ObjectId(pid)}))


# ---------------- rates ----------------
@api.get("/rates")
async def rates(date: str = "", from_cur: str = "USD", to_cur: str = "EUR",
                user: dict = Depends(get_current_user)):
    date = date or datetime.now().strftime("%Y-%m-%d")
    kur, kaynak, kur_tarihi = await tcmb.cross_rate(from_cur, to_cur, date)
    if kur is None:
        raise HTTPException(status_code=503, detail="TCMB kuru alınamadı, kuru manuel giriniz")
    return {"kur": kur, "kaynak": kaynak, "kur_tarihi": kur_tarihi,
            "from": from_cur.upper(), "to": to_cur.upper()}


# ---------------- matching ----------------
@api.get("/declarations/{did}/matches")
async def declaration_matches(did: str, user: dict = Depends(get_current_user)):
    ms = await db.matches.find({"declaration_id": did}).sort("tarih", -1).to_list(500)
    out = []
    for m in ms:
        p = await db.payments.find_one({"_id": ObjectId(m["payment_id"])})
        item = Match.from_mongo(m).model_dump()
        item["payment"] = payment_view(p) if p else None
        out.append(item)
    return out


@api.post("/matches")
async def create_match(body: MatchInput, user: dict = Depends(require("onaylayan"))):
    d = await db.declarations.find_one({"_id": ObjectId(body.declaration_id)})
    p = await db.payments.find_one({"_id": ObjectId(body.payment_id)})
    if not d or not p:
        raise HTTPException(status_code=404, detail="Beyanname veya bedel bulunamadı")
    if body.kapatilan_tutar <= 0:
        raise HTTPException(status_code=400, detail="Kapatılacak tutar sıfırdan büyük olmalı")

    d_kalan = round(d["tutar"] - d.get("kapatilan", 0), 2)
    if body.kapatilan_tutar > d_kalan + 0.01:
        raise HTTPException(status_code=400,
            detail=f"Beyanname kalan tutarı aşılamaz. Kalan: {d_kalan:,.2f} {d['doviz']}")

    if p["doviz"] == d["doviz"]:
        kur, kaynak, kur_tarihi = 1.0, "AYNI_DOVIZ", d["acilis_tarihi"][:10]
    elif body.kur:
        kur, kaynak, kur_tarihi = float(body.kur), "MANUEL", d["acilis_tarihi"][:10]
    else:
        kur, kaynak, kur_tarihi = await tcmb.cross_rate(p["doviz"], d["doviz"], d["acilis_tarihi"])
        if kur is None:
            raise HTTPException(status_code=503, detail="TCMB kuru alınamadı, kuru manuel giriniz")

    bedel_kullanilan = round(body.kapatilan_tutar / kur, 2)
    p_bakiye = round(p["tutar"] - p.get("kullanilan", 0), 2)
    if bedel_kullanilan > p_bakiye + 0.01:
        raise HTTPException(status_code=400,
            detail=f"Bedel bakiyesi yetersiz. Bakiye: {p_bakiye:,.2f} {p['doviz']} "
                   f"(gereken {bedel_kullanilan:,.2f} {p['doviz']})")

    m = Match(declaration_id=body.declaration_id, payment_id=body.payment_id,
              kapatilan_tutar=round(body.kapatilan_tutar, 2), bedel_kullanilan=bedel_kullanilan,
              kur=kur, kur_kaynak=kaynak, kullanici=user["_id"], kullanici_ad=user["name"])
    res = await db.matches.insert_one(m.to_mongo())
    await recalc_declaration(body.declaration_id)
    await recalc_payment(body.payment_id)
    await log_action(user, "Eşleştirme", "EKLE",
        f"{d['beyanname_no']} beyannamesi {body.kapatilan_tutar:,.2f} {d['doviz']} kapatıldı "
        f"({bedel_kullanilan:,.2f} {p['doviz']}, kur {kur} / {kaynak} {kur_tarihi})", str(res.inserted_id))
    return Match.from_mongo(await db.matches.find_one({"_id": res.inserted_id})).model_dump()


@api.put("/matches/{mid}")
async def update_match(mid: str, kapatilan_tutar: float, user: dict = Depends(require("onaylayan"))):
    m = await db.matches.find_one({"_id": ObjectId(mid)})
    if not m:
        raise HTTPException(status_code=404, detail="Eşleştirme bulunamadı")
    d = await db.declarations.find_one({"_id": ObjectId(m["declaration_id"])})
    p = await db.payments.find_one({"_id": ObjectId(m["payment_id"])})
    d_kalan = round(d["tutar"] - d.get("kapatilan", 0) + m["kapatilan_tutar"], 2)
    if kapatilan_tutar > d_kalan + 0.01 or kapatilan_tutar <= 0:
        raise HTTPException(status_code=400, detail=f"Geçersiz tutar. Üst limit: {d_kalan:,.2f} {d['doviz']}")
    yeni_bedel = round(kapatilan_tutar / m["kur"], 2)
    p_bakiye = round(p["tutar"] - p.get("kullanilan", 0) + m["bedel_kullanilan"], 2)
    if yeni_bedel > p_bakiye + 0.01:
        raise HTTPException(status_code=400, detail=f"Bedel bakiyesi yetersiz. Üst limit: {p_bakiye:,.2f} {p['doviz']}")
    await db.matches.update_one({"_id": ObjectId(mid)},
        {"$set": {"kapatilan_tutar": round(kapatilan_tutar, 2), "bedel_kullanilan": yeni_bedel}})
    await recalc_declaration(m["declaration_id"])
    await recalc_payment(m["payment_id"])
    await log_action(user, "Eşleştirme", "GUNCELLE",
        f"{d['beyanname_no']} eşleştirmesi {kapatilan_tutar:,.2f} {d['doviz']} olarak güncellendi", mid)
    return {"ok": True}


@api.delete("/matches/{mid}")
async def delete_match(mid: str, user: dict = Depends(require("onaylayan"))):
    m = await db.matches.find_one({"_id": ObjectId(mid)})
    if not m:
        raise HTTPException(status_code=404, detail="Eşleştirme bulunamadı")
    d = await db.declarations.find_one({"_id": ObjectId(m["declaration_id"])})
    await db.matches.delete_one({"_id": ObjectId(mid)})
    await recalc_declaration(m["declaration_id"])
    await recalc_payment(m["payment_id"])
    await log_action(user, "Eşleştirme", "GERI_AL",
        f"{d.get('beyanname_no','')} eşleştirmesi geri alındı ({m['kapatilan_tutar']:,.2f})", mid)
    return {"ok": True}


# ---------------- dashboard / audit ----------------
@api.get("/dashboard")
async def dashboard(user: dict = Depends(get_current_user)):
    decs = [declaration_view(d) for d in await db.declarations.find({}).to_list(5000)]
    pays = [payment_view(p) for p in await db.payments.find({}).to_list(5000)]

    def group(items, key, val):
        out = {}
        for i in items:
            out[i[key]] = round(out.get(i[key], 0) + i[val], 2)
        return out

    open_decs = [d for d in decs if d["durum"] != "KAPALI"]
    logs = await db.audit_logs.find({}).sort("tarih", -1).to_list(12)
    return {
        "acik": len([d for d in decs if d["durum"] == "ACIK"]),
        "kismi": len([d for d in decs if d["durum"] == "KISMI"]),
        "kapali": len([d for d in decs if d["durum"] == "KAPALI"]),
        "toplam": len(decs),
        "acik_tutar": group(open_decs, "doviz", "kalan"),
        "bedel_bakiye": group([p for p in pays if p["bakiye"] > 0.01], "doviz", "bakiye"),
        "yaklasan": sorted([d for d in decs if d["sure_durum"] == "YAKLASAN"], key=lambda x: x["kalan_gun"])[:10],
        "gecmis": sorted([d for d in decs if d["sure_durum"] == "GECMIS"], key=lambda x: x["kalan_gun"])[:10],
        "yaklasan_sayi": len([d for d in decs if d["sure_durum"] == "YAKLASAN"]),
        "gecmis_sayi": len([d for d in decs if d["sure_durum"] == "GECMIS"]),
        "son_hareketler": [{**{k: v for k, v in l.items() if k != "_id"}} for l in logs],
    }


@api.get("/audit-logs")
async def audit_logs(modul: str = "", limit: int = 200, user: dict = Depends(get_current_user)):
    query = {"modul": modul} if modul else {}
    logs = await db.audit_logs.find(query).sort("tarih", -1).to_list(limit)
    return [{k: v for k, v in l.items() if k != "_id"} for l in logs]


# ---------------- excel export ----------------
HDR_FILL = PatternFill("solid", fgColor="4338CA")


def _tr_date(s: str) -> str:
    try:
        return datetime.strptime((s or "")[:10], "%Y-%m-%d").strftime("%d.%m.%Y")
    except Exception:
        return s or ""


def _plain_style(ws, ncols):
    """Bankanın istediği sade başlık: renk/dolgu yok, sadece kalın yazı ve ince çerçeve."""
    thin = Side(style="thin", color="000000")
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="000000")
        cell.fill = PatternFill(fill_type=None)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
        ws.column_dimensions[cell.column_letter].width = 22
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"


def _style(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 20
    ws.freeze_panes = "A2"


@api.get("/export/excel")
async def export_excel(durum: str = "", user: dict = Depends(get_current_user)):
    """Bankanın istediği resmi bildirim şablonu + destek sayfaları."""
    wb = Workbook()
    ws = wb.active
    ws.title = "BANKA BİLDİRİMİ"
    headers = ["SIRA NO", "DOSYA REFERANSI", "GÜMRÜK MÜDÜRLÜĞÜ KODU", "GB NO", "GB TARİHİ",
               "GB'YE SAYILACAK TUTAR", "KULLANILACAK DTH", "KULLANILACAK ACH",
               "TCMB DEVİR ORANI", "TEŞVİK", "TAAHHÜT"]
    ws.append(headers)

    matches = await db.matches.find({}).sort("tarih", 1).to_list(5000)
    sira, toplam = 0, 0.0
    for m in matches:
        d = await db.declarations.find_one({"_id": ObjectId(m["declaration_id"])})
        p = await db.payments.find_one({"_id": ObjectId(m["payment_id"])})
        if not d or not p:
            continue
        if durum and d.get("durum") != durum:
            continue
        oran = (m["bedel_kullanilan"] / p["tutar"]) if p.get("tutar") else 0
        sira += 1
        toplam += m["kapatilan_tutar"]
        ws.append([
            sira,
            p.get("dosya_referansi", ""),
            d.get("gumruk_mudurlugu_no", ""),
            d["beyanname_no"],
            _tr_date(d.get("acilis_tarihi", "")),
            m["kapatilan_tutar"],
            p.get("dth_iban", ""),
            p.get("ach_iban") or os.environ.get("DEFAULT_ACH_IBAN", ""),
            f"%{p.get('tcmb_devir_orani', 100):g}",
            "EVET" if d.get("tesvik") else "HAYIR",
            "EVET" if d.get("taahhut") else "HAYIR",
        ])
    row = sira + 2
    ws.cell(row=row, column=1, value="TOPLAM").font = Font(bold=True)
    c = ws.cell(row=row, column=6, value=round(toplam, 2))
    c.font = Font(bold=True)
    c.number_format = "#,##0.00"
    for r in range(2, sira + 2):
        ws.cell(row=r, column=6).number_format = "#,##0.00"
        for col in (7, 8):
            ws.cell(row=r, column=col).alignment = Alignment(horizontal="left")
    _plain_style(ws, len(headers))
    ws.column_dimensions["G"].width = 32
    ws.column_dimensions["H"].width = 32

    query = {"durum": durum} if durum else {}
    decs = await db.declarations.find(query).sort("acilis_tarihi", -1).to_list(5000)
    ws1 = wb.create_sheet("Beyanname Listesi")
    h1 = ["Beyanname No", "Açılış Tarihi", "Kapanış Tarihi", "Son Kapatma Tarihi (180 gün)",
          "İthalatçı", "Gümrük Müdürlüğü No", "Döviz", "Beyanname/Fatura Tutarı",
          "Kapatılan", "Kalan", "Durum", "Süre Durumu", "IBKB Belgesi Durumu",
          "Destek Ödemesi (%3)", "Destek Durumu", "Teşvik", "Taahhüt"]
    ws1.append(h1)
    for d in decs:
        v = declaration_view(d)
        ws1.append([v["beyanname_no"], v["acilis_tarihi"], v["kapanis_tarihi"], v["son_kapatma_tarihi"],
                    v["ithalatci"], v["gumruk_mudurlugu_no"], v["doviz"], v["tutar"], v["kapatilan"],
                    v["kalan"], v["durum"], v["sure_durum"], v["ibkb_durum"],
                    v["destek_tutari"], v["destek_durum"],
                    "EVET" if v.get("tesvik") else "HAYIR", "EVET" if v.get("taahhut") else "HAYIR"])
    _style(ws1, len(h1))

    ws2 = wb.create_sheet("Eşleştirme Detayı")
    h2 = ["Beyanname No", "Bedel Gönderen", "Banka", "IBKB No", "Dosya Referansı", "Bedel Tarihi",
          "Bedel Dövizi", "Kullanılan Bedel", "DTH IBAN", "ACH IBAN", "Kur", "Kur Kaynağı",
          "Kapatılan (Beyanname Dövizi)", "Beyanname Dövizi", "İşlem Tarihi", "İşlemi Yapan"]
    ws2.append(h2)
    for m in matches:
        d = await db.declarations.find_one({"_id": ObjectId(m["declaration_id"])})
        p = await db.payments.find_one({"_id": ObjectId(m["payment_id"])})
        if not d or not p:
            continue
        ws2.append([d["beyanname_no"], p["gonderen"], p["banka"], p.get("ibkb_no", ""),
                    p.get("dosya_referansi", ""), p["tarih"], p["doviz"],
                    m["bedel_kullanilan"], p.get("dth_iban", ""),
                    p.get("ach_iban") or os.environ.get("DEFAULT_ACH_IBAN", ""),
                    m["kur"], m.get("kur_kaynak", ""), m["kapatilan_tutar"],
                    d["doviz"], m["tarih"][:19].replace("T", " "), m.get("kullanici_ad", "")])
    _style(ws2, len(h2))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    await log_action(user, "Excel", "AKTAR", "Banka bildirim Excel dosyası indirildi")
    fname = f"banka_bildirim_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})


@api.get("/reports/summary")
async def reports_summary(user: dict = Depends(get_current_user)):
    decs = [declaration_view(d) for d in await db.declarations.find({}).to_list(5000)]
    by_cur, by_imp, by_month = {}, {}, {}
    ibkb_alinmadi = destek_alinmadi = 0
    for d in decs:
        c = by_cur.setdefault(d["doviz"], {"doviz": d["doviz"], "tutar": 0, "kapatilan": 0, "kalan": 0, "adet": 0})
        c["tutar"] = round(c["tutar"] + d["tutar"], 2)
        c["kapatilan"] = round(c["kapatilan"] + d["kapatilan"], 2)
        c["kalan"] = round(c["kalan"] + d["kalan"], 2)
        c["adet"] += 1
        key = d["ithalatci"] or "-"
        k = by_imp.setdefault(key, {"ithalatci": key, "adet": 0, "kalan": 0})
        k["adet"] += 1
        k["kalan"] = round(k["kalan"] + d["kalan"], 2)
        ay = (d.get("acilis_tarihi") or "")[:7]
        m = by_month.setdefault(ay, {"ay": ay, "tutar": 0, "kapatilan": 0})
        m["tutar"] = round(m["tutar"] + d["tutar"], 2)
        m["kapatilan"] = round(m["kapatilan"] + d["kapatilan"], 2)
        if d["ibkb_durum"] == "DUZENLENMEDI":
            ibkb_alinmadi += 1
        if d["destek_durum"] == "ALINMADI":
            destek_alinmadi += 1
    return {"doviz": list(by_cur.values()),
            "ithalatci": sorted(by_imp.values(), key=lambda x: -x["adet"])[:10],
            "ibkb_alinmadi": ibkb_alinmadi, "destek_alinmadi": destek_alinmadi,
            "ay": sorted(by_month.values(), key=lambda x: x["ay"])[-12:]}


@api.get("/alerts/preview")
async def alerts_preview(user: dict = Depends(get_current_user)):
    decs = [declaration_view(d) for d in await db.declarations.find({}).to_list(5000)]
    _, counts = alerts.build_html(decs)
    return {"alicilar": alerts.recipients(), "sayilar": counts,
            "plan": "Her Pazartesi 09:00 (haftalık özet)"}


@api.post("/alerts/send")
async def alerts_send(user: dict = Depends(require("onaylayan"))):
    decs = [declaration_view(d) for d in await db.declarations.find({}).to_list(5000)]
    res = await alerts.send_alert(decs)
    if not res.get("sent"):
        raise HTTPException(status_code=502, detail=res.get("reason", "E-posta gönderilemedi"))
    await log_action(user, "Uyarı", "EPOSTA",
                     f"Haftalık uyarı e-postası gönderildi: {', '.join(res['to'])}")
    return res


async def weekly_alert_job():
    decs = [declaration_view(d) for d in await db.declarations.find({}).to_list(5000)]
    res = await alerts.send_alert(decs)
    await db.audit_logs.insert_one({
        "modul": "Uyarı", "islem": "EPOSTA",
        "aciklama": ("Haftalık uyarı e-postası otomatik gönderildi"
                     if res.get("sent") else f"Otomatik e-posta gönderilemedi: {res.get('reason')}"),
        "ref": "", "kullanici_ad": "Sistem", "kullanici_rol": "sistem", "tarih": utcnow_iso(),
    })


app.include_router(api)
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get("CORS_ORIGINS", "*").split(","),
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
async def startup():
    await db.users.create_index("email", unique=True)
    await db.declarations.create_index("beyanname_no", unique=True)
    await db.login_challenges.create_index("expires_at", expireAfterSeconds=600)
    admin_email = os.environ["ADMIN_EMAIL"].lower()
    admin_pw = os.environ["ADMIN_PASSWORD"]
    existing = await db.users.find_one({"email": admin_email})
    if not existing:
        await db.users.insert_one({"email": admin_email, "password_hash": hash_password(admin_pw),
                                   "name": "Sistem Yöneticisi", "role": "admin", "active": True,
                                   "two_factor": False, "created_at": utcnow_iso()})
    elif not verify_password(admin_pw, existing["password_hash"]):
        await db.users.update_one({"email": admin_email},
                                  {"$set": {"password_hash": hash_password(admin_pw)}})
    demo = [("ihracat@ihracat.com", "İhracat Personeli", "ihracat"),
            ("banka@ihracat.com", "Banka Personeli", "banka"),
            ("sef@ihracat.com", "Operasyon Şefi", "onaylayan"),
            ("viewer@ihracat.com", "Görüntüleyici", "goruntuleyici")]
    if os.environ.get("SEED_DEMO_USERS", "true").lower() != "true":
        demo = []
    for email, name, role in demo:
        if not await db.users.find_one({"email": email}):
            await db.users.insert_one({"email": email, "password_hash": hash_password("Test1234!"),
                                       "name": name, "role": role, "active": True,
                                       "two_factor": False, "created_at": utcnow_iso()})
    scheduler = AsyncIOScheduler(timezone="Europe/Istanbul")
    scheduler.add_job(weekly_alert_job, CronTrigger(day_of_week="mon", hour=9, minute=0),
                      id="weekly_alert", replace_existing=True)
    scheduler.start()
    app.state.scheduler = scheduler


@app.on_event("shutdown")
async def shutdown():
    client.close()
